"""
Comprehensive Excel Formula Engine for Excel Viewer Pro.
Supports 80+ functions, 2D ranges, cross-sheet references,
relative/absolute reference shifting, and IntelliSense metadata.
"""

from __future__ import annotations

import re
import math
import fnmatch
from datetime import datetime, date, timedelta
from typing import Any, Callable
from dataclasses import dataclass
from openpyxl.utils import column_index_from_string, get_column_letter


# =============================================================================
# Cell and Range Reference Regexes
# =============================================================================

# Matches A1, $A$1, A$1, $A1, Sheet1!A1, 'My Sheet'!$A$1
CELL_REF_PATTERN = re.compile(
    r"(?:(?:'([^']+)'|([A-Za-z0-9_]+))!)?(\$?)([A-Za-z]+)(\$?)(\d+)",
    re.IGNORECASE
)

# Matches A1:B10, Sheet1!A1:B10, 'My Sheet'!$A$1:$B$10
RANGE_PATTERN = re.compile(
    r"(?:(?:'([^']+)'|([A-Za-z0-9_]+))!)?(\$?[A-Za-z]+\$?\d+):(\$?[A-Za-z]+\$?\d+)",
    re.IGNORECASE
)


@dataclass(slots=True)
class CellRef:
    """Parsed cell reference with optional sheet name."""
    col: int  # 0-based
    row: int  # 0-based
    sheet: str | None = None
    col_absolute: bool = False
    row_absolute: bool = False

    @classmethod
    def from_string(cls, ref: str) -> "CellRef":
        match = CELL_REF_PATTERN.match(ref.strip())
        if not match:
            raise ValueError(f"Invalid cell reference: {ref}")
        sheet1, sheet2, col_abs, col_str, row_abs, row_str = match.groups()
        sheet = sheet1 or sheet2
        return cls(
            col=column_index_from_string(col_str.upper()) - 1,
            row=int(row_str) - 1,
            sheet=sheet,
            col_absolute=bool(col_abs),
            row_absolute=bool(row_abs)
        )

    def to_string(self) -> str:
        col_prefix = "$" if self.col_absolute else ""
        row_prefix = "$" if self.row_absolute else ""
        col_letter = get_column_letter(self.col + 1)
        ref = f"{col_prefix}{col_letter}{row_prefix}{self.row + 1}"
        if self.sheet:
            if " " in self.sheet or "'" in self.sheet:
                return f"'{self.sheet}'!{ref}"
            return f"{self.sheet}!{ref}"
        return ref


def parse_range(range_str: str) -> list[CellRef]:
    """Parse range like 'A1:B3' or 'Sheet1!A1:B3' into list of CellRefs."""
    match = RANGE_PATTERN.match(range_str.strip())
    if not match:
        raise ValueError(f"Invalid range: {range_str}")

    sheet1, sheet2, start_str, end_str = match.groups()
    sheet = sheet1 or sheet2

    start = CellRef.from_string(start_str)
    end = CellRef.from_string(end_str)

    min_row = min(start.row, end.row)
    max_row = max(start.row, end.row)
    min_col = min(start.col, end.col)
    max_col = max(start.col, end.col)

    cells = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cells.append(CellRef(col=c, row=r, sheet=sheet))
    return cells


def shift_formula_references(formula: str, row_delta: int, col_delta: int) -> str:
    """
    Shifts relative references in an Excel formula by row_delta and col_delta.
    Preserves absolute references ($A$1, A$1, $A1).
    """
    if not formula or not formula.startswith("="):
        return formula

    def replace_cell_match(m: re.Match) -> str:
        full_match = m.group(0)
        sheet1, sheet2, col_abs, col_str, row_abs, row_str = m.groups()
        sheet = sheet1 or sheet2

        col_idx = column_index_from_string(col_str.upper()) - 1
        row_idx = int(row_str) - 1

        if not col_abs:
            col_idx = max(0, col_idx + col_delta)
        if not row_abs:
            row_idx = max(0, row_idx + row_delta)

        new_ref = f"{'$' if col_abs else ''}{get_column_letter(col_idx + 1)}{'$' if row_abs else ''}{row_idx + 1}"
        if sheet:
            if " " in sheet or "'" in sheet:
                return f"'{sheet}'!{new_ref}"
            return f"{sheet}!{new_ref}"
        return new_ref

    # We do a replacement on all CELL_REF_PATTERNs
    return CELL_REF_PATTERN.sub(replace_cell_match, formula)


# =============================================================================
# Function Directory and Metadata (for IntelliSense and Wizard)
# =============================================================================

FUNCTION_METADATA: dict[str, dict[str, Any]] = {
    # Math & Trig
    "SUM": {"cat": "Math & Trig", "syntax": "SUM(number1, [number2], ...)", "desc": "Adds all numbers in a range of cells."},
    "SUMIF": {"cat": "Math & Trig", "syntax": "SUMIF(range, criteria, [sum_range])", "desc": "Adds cells specified by a given condition."},
    "SUMIFS": {"cat": "Math & Trig", "syntax": "SUMIFS(sum_range, criteria_range1, criteria1, ...)", "desc": "Adds cells specified by multiple conditions."},
    "PRODUCT": {"cat": "Math & Trig", "syntax": "PRODUCT(number1, [number2], ...)", "desc": "Multiplies all numbers given as arguments."},
    "SUMPRODUCT": {"cat": "Math & Trig", "syntax": "SUMPRODUCT(array1, [array2], ...)", "desc": "Returns the sum of the products of corresponding array components."},
    "ABS": {"cat": "Math & Trig", "syntax": "ABS(number)", "desc": "Returns the absolute value of a number."},
    "ROUND": {"cat": "Math & Trig", "syntax": "ROUND(number, num_digits)", "desc": "Rounds a number to a specified number of digits."},
    "ROUNDUP": {"cat": "Math & Trig", "syntax": "ROUNDUP(number, num_digits)", "desc": "Rounds a number up, away from zero."},
    "ROUNDDOWN": {"cat": "Math & Trig", "syntax": "ROUNDDOWN(number, num_digits)", "desc": "Rounds a number down, toward zero."},
    "INT": {"cat": "Math & Trig", "syntax": "INT(number)", "desc": "Rounds a number down to the nearest integer."},
    "TRUNC": {"cat": "Math & Trig", "syntax": "TRUNC(number, [num_digits])", "desc": "Truncates a number to an integer or specified digits."},
    "MOD": {"cat": "Math & Trig", "syntax": "MOD(number, divisor)", "desc": "Returns the remainder after number is divided by divisor."},
    "POWER": {"cat": "Math & Trig", "syntax": "POWER(number, power)", "desc": "Returns the result of a number raised to a power."},
    "SQRT": {"cat": "Math & Trig", "syntax": "SQRT(number)", "desc": "Returns the positive square root of a number."},
    "PI": {"cat": "Math & Trig", "syntax": "PI()", "desc": "Returns the value of Pi (3.14159265...)."},
    "RAND": {"cat": "Math & Trig", "syntax": "RAND()", "desc": "Returns a random number between 0 and 1."},
    "RANDBETWEEN": {"cat": "Math & Trig", "syntax": "RANDBETWEEN(bottom, top)", "desc": "Returns a random integer between the numbers you specify."},
    "CEILING": {"cat": "Math & Trig", "syntax": "CEILING(number, [significance])", "desc": "Rounds a number up to the nearest multiple of significance."},
    "FLOOR": {"cat": "Math & Trig", "syntax": "FLOOR(number, [significance])", "desc": "Rounds a number down to the nearest multiple of significance."},
    "SIGN": {"cat": "Math & Trig", "syntax": "SIGN(number)", "desc": "Returns 1 if positive, -1 if negative, 0 if zero."},
    "SIN": {"cat": "Math & Trig", "syntax": "SIN(number)", "desc": "Returns the sine of the given angle in radians."},
    "COS": {"cat": "Math & Trig", "syntax": "COS(number)", "desc": "Returns the cosine of the given angle in radians."},
    "TAN": {"cat": "Math & Trig", "syntax": "TAN(number)", "desc": "Returns the tangent of the given angle in radians."},
    "ASIN": {"cat": "Math & Trig", "syntax": "ASIN(number)", "desc": "Returns the arcsine of a number in radians."},
    "ACOS": {"cat": "Math & Trig", "syntax": "ACOS(number)", "desc": "Returns the arccosine of a number in radians."},
    "ATAN": {"cat": "Math & Trig", "syntax": "ATAN(number)", "desc": "Returns the arctangent of a number in radians."},
    "DEGREES": {"cat": "Math & Trig", "syntax": "DEGREES(angle)", "desc": "Converts radians into degrees."},
    "RADIANS": {"cat": "Math & Trig", "syntax": "RADIANS(angle)", "desc": "Converts degrees into radians."},
    "EXP": {"cat": "Math & Trig", "syntax": "EXP(number)", "desc": "Returns e raised to the power of number."},
    "LN": {"cat": "Math & Trig", "syntax": "LN(number)", "desc": "Returns the natural logarithm of a number."},
    "LOG": {"cat": "Math & Trig", "syntax": "LOG(number, [base])", "desc": "Returns the logarithm of a number to a specified base."},
    "LOG10": {"cat": "Math & Trig", "syntax": "LOG10(number)", "desc": "Returns the base-10 logarithm of a number."},
    "FACT": {"cat": "Math & Trig", "syntax": "FACT(number)", "desc": "Returns the factorial of a number."},

    # Statistical
    "AVERAGE": {"cat": "Statistical", "syntax": "AVERAGE(number1, [number2], ...)", "desc": "Returns the average (arithmetic mean) of its arguments."},
    "AVERAGEA": {"cat": "Statistical", "syntax": "AVERAGEA(value1, [value2], ...)", "desc": "Returns the average of its arguments, evaluating text and FALSE as 0."},
    "AVERAGEIF": {"cat": "Statistical", "syntax": "AVERAGEIF(range, criteria, [average_range])", "desc": "Returns the average of cells that meet a given condition."},
    "AVERAGEIFS": {"cat": "Statistical", "syntax": "AVERAGEIFS(avg_range, criteria_range1, criteria1, ...)", "desc": "Returns the average of cells that meet multiple criteria."},
    "COUNT": {"cat": "Statistical", "syntax": "COUNT(value1, [value2], ...)", "desc": "Counts how many numbers are in the list of arguments."},
    "COUNTA": {"cat": "Statistical", "syntax": "COUNTA(value1, [value2], ...)", "desc": "Counts how many values in the list of arguments are not empty."},
    "COUNTBLANK": {"cat": "Statistical", "syntax": "COUNTBLANK(range)", "desc": "Counts empty cells in a specified range."},
    "COUNTIF": {"cat": "Statistical", "syntax": "COUNTIF(range, criteria)", "desc": "Counts the number of cells within a range that meet the given criteria."},
    "COUNTIFS": {"cat": "Statistical", "syntax": "COUNTIFS(criteria_range1, criteria1, ...)", "desc": "Counts the number of cells within ranges that meet multiple criteria."},
    "MIN": {"cat": "Statistical", "syntax": "MIN(number1, [number2], ...)", "desc": "Returns the minimum value in a list of arguments."},
    "MAX": {"cat": "Statistical", "syntax": "MAX(number1, [number2], ...)", "desc": "Returns the maximum value in a list of arguments."},
    "MINIFS": {"cat": "Statistical", "syntax": "MINIFS(min_range, criteria_range1, criteria1, ...)", "desc": "Returns the minimum value among cells specified by a set of criteria."},
    "MAXIFS": {"cat": "Statistical", "syntax": "MAXIFS(max_range, criteria_range1, criteria1, ...)", "desc": "Returns the maximum value among cells specified by a set of criteria."},
    "MEDIAN": {"cat": "Statistical", "syntax": "MEDIAN(number1, [number2], ...)", "desc": "Returns the median of the given numbers."},
    "MODE": {"cat": "Statistical", "syntax": "MODE(number1, [number2], ...)", "desc": "Returns the most frequently occurring value in a range."},
    "STDEV": {"cat": "Statistical", "syntax": "STDEV(number1, [number2], ...)", "desc": "Estimates standard deviation based on a sample."},
    "STDEVP": {"cat": "Statistical", "syntax": "STDEVP(number1, [number2], ...)", "desc": "Calculates standard deviation based on the entire population."},
    "VAR": {"cat": "Statistical", "syntax": "VAR(number1, [number2], ...)", "desc": "Estimates variance based on a sample."},
    "VARP": {"cat": "Statistical", "syntax": "VARP(number1, [number2], ...)", "desc": "Calculates variance based on the entire population."},
    "LARGE": {"cat": "Statistical", "syntax": "LARGE(array, k)", "desc": "Returns the k-th largest value in a data set."},
    "SMALL": {"cat": "Statistical", "syntax": "SMALL(array, k)", "desc": "Returns the k-th smallest value in a data set."},
    "RANK": {"cat": "Statistical", "syntax": "RANK(number, ref, [order])", "desc": "Returns the rank of a number in a list of numbers."},

    # Lookup & Reference
    "VLOOKUP": {"cat": "Lookup & Reference", "syntax": "VLOOKUP(lookup_value, table_array, col_index_num, [range_lookup])", "desc": "Looks for a value in the leftmost column of a table and returns a value in the same row from a specified column."},
    "HLOOKUP": {"cat": "Lookup & Reference", "syntax": "HLOOKUP(lookup_value, table_array, row_index_num, [range_lookup])", "desc": "Looks for a value in the top row of a table and returns a value in the same column from a specified row."},
    "XLOOKUP": {"cat": "Lookup & Reference", "syntax": "XLOOKUP(lookup_value, lookup_array, return_array, [if_not_found], [match_mode])", "desc": "Modern replacement for VLOOKUP, searching in any array and returning corresponding item from another."},
    "INDEX": {"cat": "Lookup & Reference", "syntax": "INDEX(array, row_num, [col_num])", "desc": "Returns the value of a specified cell or array of cells within a table or range."},
    "MATCH": {"cat": "Lookup & Reference", "syntax": "MATCH(lookup_value, lookup_array, [match_type])", "desc": "Returns the relative position of an item in an array that matches a specified value."},
    "LOOKUP": {"cat": "Lookup & Reference", "syntax": "LOOKUP(lookup_value, lookup_vector, [result_vector])", "desc": "Looks up a value either from a one-row or one-column range."},
    "CHOOSE": {"cat": "Lookup & Reference", "syntax": "CHOOSE(index_num, value1, [value2], ...)", "desc": "Uses index_num to return a value from the list of value arguments."},
    "ROW": {"cat": "Lookup & Reference", "syntax": "ROW([reference])", "desc": "Returns the row number of a reference."},
    "COLUMN": {"cat": "Lookup & Reference", "syntax": "COLUMN([reference])", "desc": "Returns the column number of a reference."},
    "ROWS": {"cat": "Lookup & Reference", "syntax": "ROWS(array)", "desc": "Returns the number of rows in a reference or array."},
    "COLUMNS": {"cat": "Lookup & Reference", "syntax": "COLUMNS(array)", "desc": "Returns the number of columns in a reference or array."},

    # Logical
    "IF": {"cat": "Logical", "syntax": "IF(logical_test, value_if_true, [value_if_false])", "desc": "Checks whether a condition is met, and returns one value if TRUE, and another if FALSE."},
    "IFS": {"cat": "Logical", "syntax": "IFS(logical_test1, value1, [logical_test2, value2], ...)", "desc": "Checks whether one or more conditions are met and returns a value that corresponds to the first TRUE condition."},
    "AND": {"cat": "Logical", "syntax": "AND(logical1, [logical2], ...)", "desc": "Returns TRUE if all its arguments are TRUE."},
    "OR": {"cat": "Logical", "syntax": "OR(logical1, [logical2], ...)", "desc": "Returns TRUE if any argument is TRUE."},
    "NOT": {"cat": "Logical", "syntax": "NOT(logical)", "desc": "Reverses the logic of its argument."},
    "XOR": {"cat": "Logical", "syntax": "XOR(logical1, [logical2], ...)", "desc": "Returns a logical exclusive OR of all arguments."},
    "IFERROR": {"cat": "Logical", "syntax": "IFERROR(value, value_if_error)", "desc": "Returns value_if_error if expression is an error and the value of the expression itself otherwise."},
    "IFNA": {"cat": "Logical", "syntax": "IFNA(value, value_if_na)", "desc": "Returns value_if_na if expression resolves to #N/A."},
    "TRUE": {"cat": "Logical", "syntax": "TRUE()", "desc": "Returns the logical value TRUE."},
    "FALSE": {"cat": "Logical", "syntax": "FALSE()", "desc": "Returns the logical value FALSE."},
    "SWITCH": {"cat": "Logical", "syntax": "SWITCH(expression, val1, result1, [val2, result2], ..., [default])", "desc": "Evaluates an expression against a list of values and returns the result corresponding to the first matching value."},

    # Text
    "CONCAT": {"cat": "Text", "syntax": "CONCAT(text1, [text2], ...)", "desc": "Combines text from multiple ranges and/or strings."},
    "CONCATENATE": {"cat": "Text", "syntax": "CONCATENATE(text1, [text2], ...)", "desc": "Joins several text strings into one text string."},
    "TEXTJOIN": {"cat": "Text", "syntax": "TEXTJOIN(delimiter, ignore_empty, text1, [text2], ...)", "desc": "Combines text from multiple ranges with a specified delimiter."},
    "LEFT": {"cat": "Text", "syntax": "LEFT(text, [num_chars])", "desc": "Returns the specified number of characters from the start of a text string."},
    "RIGHT": {"cat": "Text", "syntax": "RIGHT(text, [num_chars])", "desc": "Returns the specified number of characters from the end of a text string."},
    "MID": {"cat": "Text", "syntax": "MID(text, start_num, num_chars)", "desc": "Returns a specific number of characters from a text string, starting at the position you specify."},
    "LEN": {"cat": "Text", "syntax": "LEN(text)", "desc": "Returns the number of characters in a text string."},
    "TRIM": {"cat": "Text", "syntax": "TRIM(text)", "desc": "Removes all spaces from text except for single spaces between words."},
    "UPPER": {"cat": "Text", "syntax": "UPPER(text)", "desc": "Converts text to uppercase."},
    "LOWER": {"cat": "Text", "syntax": "LOWER(text)", "desc": "Converts text to lowercase."},
    "PROPER": {"cat": "Text", "syntax": "PROPER(text)", "desc": "Capitalizes the first letter in each word of a text value."},
    "EXACT": {"cat": "Text", "syntax": "EXACT(text1, text2)", "desc": "Checks whether two text strings are exactly the same (case-sensitive)."},
    "FIND": {"cat": "Text", "syntax": "FIND(find_text, within_text, [start_num])", "desc": "Finds one text value within another (case-sensitive)."},
    "SEARCH": {"cat": "Text", "syntax": "SEARCH(find_text, within_text, [start_num])", "desc": "Finds one text value within another (not case-sensitive)."},
    "REPLACE": {"cat": "Text", "syntax": "REPLACE(old_text, start_num, num_chars, new_text)", "desc": "Replaces characters within text."},
    "SUBSTITUTE": {"cat": "Text", "syntax": "SUBSTITUTE(text, old_text, new_text, [instance_num])", "desc": "Substitutes new text for old text in a text string."},
    "REPT": {"cat": "Text", "syntax": "REPT(text, number_times)", "desc": "Repeats text a given number of times."},
    "TEXT": {"cat": "Text", "syntax": "TEXT(value, format_text)", "desc": "Formats a number and converts it to text."},
    "VALUE": {"cat": "Text", "syntax": "VALUE(text)", "desc": "Converts a text string that represents a number to a number."},
    "CHAR": {"cat": "Text", "syntax": "CHAR(number)", "desc": "Returns the character specified by the code number."},
    "CODE": {"cat": "Text", "syntax": "CODE(text)", "desc": "Returns a numeric code for the first character in a text string."},
    "CLEAN": {"cat": "Text", "syntax": "CLEAN(text)", "desc": "Removes all nonprintable characters from text."},

    # Date & Time
    "TODAY": {"cat": "Date & Time", "syntax": "TODAY()", "desc": "Returns the current date."},
    "NOW": {"cat": "Date & Time", "syntax": "NOW()", "desc": "Returns the current date and time."},
    "DATE": {"cat": "Date & Time", "syntax": "DATE(year, month, day)", "desc": "Returns the serial number that represents a particular date."},
    "TIME": {"cat": "Date & Time", "syntax": "TIME(hour, minute, second)", "desc": "Returns the decimal number for a particular time."},
    "YEAR": {"cat": "Date & Time", "syntax": "YEAR(serial_number)", "desc": "Returns the year corresponding to a date."},
    "MONTH": {"cat": "Date & Time", "syntax": "MONTH(serial_number)", "desc": "Returns the month corresponding to a date."},
    "DAY": {"cat": "Date & Time", "syntax": "DAY(serial_number)", "desc": "Returns the day of the month corresponding to a date."},
    "HOUR": {"cat": "Date & Time", "syntax": "HOUR(serial_number)", "desc": "Returns the hour corresponding to a time."},
    "MINUTE": {"cat": "Date & Time", "syntax": "MINUTE(serial_number)", "desc": "Returns the minute corresponding to a time."},
    "SECOND": {"cat": "Date & Time", "syntax": "SECOND(serial_number)", "desc": "Returns the second corresponding to a time."},
    "WEEKDAY": {"cat": "Date & Time", "syntax": "WEEKDAY(serial_number, [return_type])", "desc": "Returns the day of the week corresponding to a date."},
    "DAYS": {"cat": "Date & Time", "syntax": "DAYS(end_date, start_date)", "desc": "Returns the number of days between two dates."},
    "DATEDIF": {"cat": "Date & Time", "syntax": "DATEDIF(start_date, end_date, unit)", "desc": "Calculates the number of days, months, or years between two dates."},
    "EDATE": {"cat": "Date & Time", "syntax": "EDATE(start_date, months)", "desc": "Returns the serial number of the date that is the indicated number of months before or after the start_date."},
    "EOMONTH": {"cat": "Date & Time", "syntax": "EOMONTH(start_date, months)", "desc": "Returns the serial number of the last day of the month before or after a specified number of months."},

    # Financial
    "PMT": {"cat": "Financial", "syntax": "PMT(rate, nper, pv, [fv], [type])", "desc": "Calculates the payment for a loan based on constant payments and a constant interest rate."},
    "PV": {"cat": "Financial", "syntax": "PV(rate, nper, pmt, [fv], [type])", "desc": "Returns the present value of an investment."},
    "FV": {"cat": "Financial", "syntax": "FV(rate, nper, pmt, [pv], [type])", "desc": "Returns the future value of an investment based on constant payments and a constant interest rate."},
    "NPER": {"cat": "Financial", "syntax": "NPER(rate, pmt, pv, [fv], [type])", "desc": "Returns the number of periods for an investment based on periodic, constant payments and a constant interest rate."},
    "RATE": {"cat": "Financial", "syntax": "RATE(nper, pmt, pv, [fv], [type], [guess])", "desc": "Returns the interest rate per period of an annuity."},

    # Information
    "ISBLANK": {"cat": "Information", "syntax": "ISBLANK(value)", "desc": "Returns TRUE if the value is blank."},
    "ISNUMBER": {"cat": "Information", "syntax": "ISNUMBER(value)", "desc": "Returns TRUE if the value is a number."},
    "ISTEXT": {"cat": "Information", "syntax": "ISTEXT(value)", "desc": "Returns TRUE if the value is text."},
    "ISNONTEXT": {"cat": "Information", "syntax": "ISNONTEXT(value)", "desc": "Returns TRUE if the value is not text."},
    "ISLOGICAL": {"cat": "Information", "syntax": "ISLOGICAL(value)", "desc": "Returns TRUE if the value is a logical value (TRUE or FALSE)."},
    "ISERROR": {"cat": "Information", "syntax": "ISERROR(value)", "desc": "Returns TRUE if the value is any error value (#N/A, #VALUE!, #REF!, #DIV/0!, #NUM!, #NAME?, or #NULL!)."},
    "ISERR": {"cat": "Information", "syntax": "ISERR(value)", "desc": "Returns TRUE if the value is any error value except #N/A."},
    "ISNA": {"cat": "Information", "syntax": "ISNA(value)", "desc": "Returns TRUE if the value is the #N/A error value."},
    "TYPE": {"cat": "Information", "syntax": "TYPE(value)", "desc": "Returns a number indicating the data type of a value (1=number, 2=text, 4=logical, 16=error, 64=array)."},
    "N": {"cat": "Information", "syntax": "N(value)", "desc": "Returns a value converted to a number."}
}


# =============================================================================
# Formula Engine
# =============================================================================

class FormulaEngine:
    """Excel-like formula parser and evaluator with 80+ functions and multi-sheet support."""

    def __init__(self, get_cell_value_callback: Callable[[int, int, str | None], Any]) -> None:
        """
        get_cell_value_callback: func(row, col, sheet_name=None) -> Any
        """
        self._get_val = get_cell_value_callback
        self._functions = self._build_function_map()

    def evaluate(self, formula: str, current_sheet: str | None = None) -> Any:
        """Evaluate formula string, starting with '='."""
        if not formula or not isinstance(formula, str) or not formula.startswith("="):
            return formula

        expr = formula[1:].strip()
        if not expr:
            return ""

        try:
            res = self._eval_expr(expr, current_sheet)
            if isinstance(res, float):
                if math.isnan(res):
                    return "#NUM!"
                if math.isinf(res):
                    return "#DIV/0!"
                if res == int(res):
                    return int(res)
                return round(res, 10)
            return res
        except ZeroDivisionError:
            return "#DIV/0!"
        except ValueError as e:
            return f"#VALUE!"
        except Exception as e:
            return f"#ERROR: {e}"

    def _eval_expr(self, expr: str, current_sheet: str | None) -> Any:
        expr = expr.strip()
        if not expr:
            return ""

        # String literal "text"
        if expr.startswith('"') and expr.endswith('"') and len(expr) >= 2:
            # Handle escaped quotes ""
            return expr[1:-1].replace('""', '"')

        # Numbers
        try:
            if "." in expr or "e" in expr.lower():
                return float(expr)
            return int(expr)
        except ValueError:
            pass

        # Boolean
        if expr.upper() == "TRUE":
            return True
        if expr.upper() == "FALSE":
            return False

        # Functions FUNC(...)
        func_match = re.match(r"^([A-Z0-9_\.]+)\s*\((.*)\)$", expr, re.IGNORECASE | re.DOTALL)
        if func_match:
            fn_name = func_match.group(1).upper()
            args_str = func_match.group(2)
            if fn_name in self._functions:
                return self._call_func(fn_name, args_str, current_sheet)

        # Comparisons (=, <>, <=, >=, <, >)
        for op in ["<>", "<=", ">=", "<", ">", "="]:
            parts = self._split_top_level(expr, op)
            if len(parts) == 2:
                left = self._eval_expr(parts[0], current_sheet)
                right = self._eval_expr(parts[1], current_sheet)
                return self._compare(left, right, op)

        # String concatenation (&)
        parts = self._split_top_level(expr, "&")
        if len(parts) == 2:
            left = self._eval_expr(parts[0], current_sheet)
            right = self._eval_expr(parts[1], current_sheet)
            return f"{left if left is not None else ''}{right if right is not None else ''}"

        # Add / Sub (+, -)
        for op in ["+", "-"]:
            parts = self._split_top_level(expr, op)
            if len(parts) == 2 and parts[0].strip() != "":
                left = self._to_number(self._eval_expr(parts[0], current_sheet))
                right = self._to_number(self._eval_expr(parts[1], current_sheet))
                return left + right if op == "+" else left - right

        # Mult / Div / Mod (*, /, %)
        for op in ["*", "/", "%"]:
            parts = self._split_top_level(expr, op)
            if len(parts) == 2:
                left = self._to_number(self._eval_expr(parts[0], current_sheet))
                right = self._to_number(self._eval_expr(parts[1], current_sheet))
                if op == "*":
                    return left * right
                elif op == "/":
                    if right == 0:
                        return "#DIV/0!"
                    return left / right
                else:
                    return left % right

        # Power (^)
        parts = self._split_top_level(expr, "^")
        if len(parts) == 2:
            base = self._to_number(self._eval_expr(parts[0], current_sheet))
            pwr = self._to_number(self._eval_expr(parts[1], current_sheet))
            return math.pow(base, pwr)

        # Parentheses
        if expr.startswith("(") and expr.endswith(")"):
            inner = expr[1:-1].strip()
            # check matching parentheses
            if self._is_balanced(inner):
                return self._eval_expr(inner, current_sheet)

        # Single Cell Reference
        if CELL_REF_PATTERN.match(expr):
            ref = CellRef.from_string(expr)
            sheet = ref.sheet or current_sheet
            return self._get_val(ref.row, ref.col, sheet)

        return expr

    def _split_top_level(self, expr: str, op: str) -> list[str]:
        """Split expression by operator only at top parenthesis and quote depth."""
        depth = 0
        in_quote = False
        op_len = len(op)

        # Scan right to left for left-associative operators
        for i in range(len(expr) - op_len, -1, -1):
            c = expr[i]
            if c == '"':
                in_quote = not in_quote
            elif not in_quote:
                if c == ")":
                    depth += 1
                elif c == "(":
                    depth -= 1
                elif depth == 0:
                    if expr[i:i + op_len] == op:
                        # Make sure not inside comparison tokens like <= or <>
                        if op in "<>=" and i > 0 and expr[i - 1] in "<>=":
                            continue
                        if op in "<>" and i + 1 < len(expr) and expr[i + 1] in "<>=":
                            continue
                        return [expr[:i], expr[i + op_len:]]
        return [expr]

    def _is_balanced(self, text: str) -> bool:
        depth = 0
        in_quote = False
        for c in text:
            if c == '"':
                in_quote = not in_quote
            elif not in_quote:
                if c == "(":
                    depth += 1
                elif c == ")":
                    depth -= 1
                    if depth < 0:
                        return False
        return depth == 0

    def _split_args(self, args_str: str) -> list[str]:
        """Split function arguments respecting nested commas inside parens/quotes."""
        args = []
        cur = []
        depth = 0
        in_quote = False
        for c in args_str:
            if c == '"':
                in_quote = not in_quote
                cur.append(c)
            elif not in_quote:
                if c == "(":
                    depth += 1
                    cur.append(c)
                elif c == ")":
                    depth -= 1
                    cur.append(c)
                elif c == "," and depth == 0:
                    args.append("".join(cur).strip())
                    cur.clear()
                else:
                    cur.append(c)
            else:
                cur.append(c)
        if cur:
            args.append("".join(cur).strip())
        return args

    def _expand_arg_raw(self, arg_str: str, current_sheet: str | None) -> list[Any]:
        """Expand argument string: if range, returns flat list of cell values."""
        arg_str = arg_str.strip()
        if not arg_str:
            return []

        # Check if it's a range like A1:B10 or Sheet1!A1:B10
        if RANGE_PATTERN.match(arg_str):
            cells = parse_range(arg_str)
            vals = []
            for cell in cells:
                sheet = cell.sheet or current_sheet
                vals.append(self._get_val(cell.row, cell.col, sheet))
            return vals

        # Check single cell ref
        if CELL_REF_PATTERN.match(arg_str):
            cell = CellRef.from_string(arg_str)
            sheet = cell.sheet or current_sheet
            return [self._get_val(cell.row, cell.col, sheet)]

        # Evaluate expression
        val = self._eval_expr(arg_str, current_sheet)
        return [val]

    def _call_func(self, name: str, args_str: str, current_sheet: str | None) -> Any:
        fn = self._functions.get(name)
        if not fn:
            return f"#NAME? ({name})"

        raw_args = self._split_args(args_str)

        # Special functions needing raw expressions or ranges
        if name in ("IF", "IFS", "IFERROR", "IFNA", "SWITCH", "CHOOSE"):
            return fn(raw_args, current_sheet, self)

        if name in ("VLOOKUP", "HLOOKUP", "XLOOKUP", "INDEX", "MATCH", "COUNTIF", "COUNTIFS", "SUMIF", "SUMIFS", "AVERAGEIF", "AVERAGEIFS", "MINIFS", "MAXIFS", "LOOKUP", "COUNTBLANK", "ROWS", "COLUMNS"):
            return fn(raw_args, current_sheet, self)

        # General functions: expand ranges and evaluate
        expanded = []
        for arg in raw_args:
            vals = self._expand_arg_raw(arg, current_sheet)
            expanded.extend(vals)

        return fn(expanded)

    # -------------------------------------------------------------------------
    # Type Conversion Helpers
    # -------------------------------------------------------------------------

    def _to_number(self, val: Any) -> float:
        if val is None or val == "":
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        if isinstance(val, bool):
            return 1.0 if val else 0.0
        try:
            s = str(val).replace(",", ".").replace(" ", "").replace("₽", "").replace("$", "").replace("€", "").replace("%", "")
            if "%" in str(val):
                return float(s) / 100.0
            return float(s)
        except (ValueError, TypeError):
            return 0.0

    def _to_bool(self, val: Any) -> bool:
        if isinstance(val, bool):
            return val
        if isinstance(val, (int, float)):
            return val != 0
        if isinstance(val, str):
            u = val.strip().upper()
            if u in ("TRUE", "1", "YES", "ДА", "T"):
                return True
            if u in ("FALSE", "0", "NO", "НЕТ", "F", ""):
                return False
        return bool(val)

    def _compare(self, left: Any, right: Any, op: str) -> bool:
        # Numeric comparison if both can be numbers
        try:
            l_num = self._to_number(left)
            r_num = self._to_number(right)
            if isinstance(left, (int, float, bool)) or isinstance(right, (int, float, bool)) or (
                str(left).replace(".", "").replace("-", "").isdigit() and str(right).replace(".", "").replace("-", "").isdigit()
            ):
                if op == "=": return l_num == r_num
                if op == "<>": return l_num != r_num
                if op == "<": return l_num < r_num
                if op == ">": return l_num > r_num
                if op == "<=": return l_num <= r_num
                if op == ">=": return l_num >= r_num
        except Exception:
            pass

        # String comparison
        l_str = str(left if left is not None else "")
        r_str = str(right if right is not None else "")
        if op == "=": return l_str.lower() == r_str.lower()
        if op == "<>": return l_str.lower() != r_str.lower()
        if op == "<": return l_str < r_str
        if op == ">": return l_str > r_str
        if op == "<=": return l_str <= r_str
        if op == ">=": return l_str >= r_str
        return False

    def _match_criteria(self, val: Any, criteria: Any) -> bool:
        """Matches a value against criteria string (e.g. '>10', '<=5', 'Apples', '*test*')."""
        c_str = str(criteria).strip()
        if not c_str:
            return val is None or val == ""

        for op in [">=", "<=", "<>", ">", "<", "="]:
            if c_str.startswith(op):
                target = c_str[len(op):].strip()
                try:
                    t_num = float(target)
                    v_num = self._to_number(val)
                    if op == ">=": return v_num >= t_num
                    if op == "<=": return v_num <= t_num
                    if op == ">": return v_num > t_num
                    if op == "<": return v_num < t_num
                    if op == "=": return v_num == t_num
                    if op == "<>": return v_num != t_num
                except ValueError:
                    return self._compare(val, target, op)

        # Exact match or wildcard
        v_str = str(val if val is not None else "").lower()
        t_str = c_str.lower()
        if "*" in t_str or "?" in t_str:
            return fnmatch.fnmatch(v_str, t_str)
        return v_str == t_str

    # -------------------------------------------------------------------------
    # Build Function Registry
    # -------------------------------------------------------------------------

    def _build_function_map(self) -> dict[str, Callable]:
        m: dict[str, Callable] = {}

        # Math
        m["SUM"] = lambda args: sum(self._to_number(a) for a in args if a is not None and a != "")
        m["PRODUCT"] = lambda args: math.prod(self._to_number(a) for a in args if a is not None and a != "") if args else 0
        m["ABS"] = lambda args: abs(self._to_number(args[0])) if args else 0
        m["ROUND"] = lambda args: round(self._to_number(args[0]), int(self._to_number(args[1]))) if len(args) > 1 else round(self._to_number(args[0]))
        m["ROUNDUP"] = lambda args: math.ceil(self._to_number(args[0]) * (10 ** int(self._to_number(args[1])))) / (10 ** int(self._to_number(args[1]))) if len(args) > 1 else math.ceil(self._to_number(args[0]))
        m["ROUNDDOWN"] = lambda args: math.floor(self._to_number(args[0]) * (10 ** int(self._to_number(args[1])))) / (10 ** int(self._to_number(args[1]))) if len(args) > 1 else math.floor(self._to_number(args[0]))
        m["INT"] = lambda args: int(math.floor(self._to_number(args[0]))) if args else 0
        m["TRUNC"] = lambda args: int(self._to_number(args[0])) if args else 0
        m["MOD"] = lambda args: self._to_number(args[0]) % self._to_number(args[1]) if len(args) > 1 and self._to_number(args[1]) != 0 else 0
        m["POWER"] = lambda args: math.pow(self._to_number(args[0]), self._to_number(args[1])) if len(args) > 1 else 0
        m["SQRT"] = lambda args: math.sqrt(self._to_number(args[0])) if args and self._to_number(args[0]) >= 0 else "#NUM!"
        m["PI"] = lambda args: math.pi
        m["RAND"] = lambda args: __import__("random").random()
        m["RANDBETWEEN"] = lambda args: __import__("random").randint(int(self._to_number(args[0])), int(self._to_number(args[1]))) if len(args) > 1 else 0
        m["CEILING"] = lambda args: math.ceil(self._to_number(args[0])) if args else 0
        m["FLOOR"] = lambda args: math.floor(self._to_number(args[0])) if args else 0
        m["SIGN"] = lambda args: 1 if self._to_number(args[0]) > 0 else (-1 if self._to_number(args[0]) < 0 else 0) if args else 0
        m["EXP"] = lambda args: math.exp(self._to_number(args[0])) if args else 1
        m["LN"] = lambda args: math.log(self._to_number(args[0])) if args and self._to_number(args[0]) > 0 else "#NUM!"
        m["LOG"] = lambda args: math.log(self._to_number(args[0]), self._to_number(args[1])) if len(args) > 1 and self._to_number(args[0]) > 0 else (math.log10(self._to_number(args[0])) if args and self._to_number(args[0]) > 0 else "#NUM!")
        m["LOG10"] = lambda args: math.log10(self._to_number(args[0])) if args and self._to_number(args[0]) > 0 else "#NUM!"
        m["SIN"] = lambda args: math.sin(self._to_number(args[0])) if args else 0
        m["COS"] = lambda args: math.cos(self._to_number(args[0])) if args else 0
        m["TAN"] = lambda args: math.tan(self._to_number(args[0])) if args else 0
        m["ASIN"] = lambda args: math.asin(self._to_number(args[0])) if args and -1 <= self._to_number(args[0]) <= 1 else "#NUM!"
        m["ACOS"] = lambda args: math.acos(self._to_number(args[0])) if args and -1 <= self._to_number(args[0]) <= 1 else "#NUM!"
        m["ATAN"] = lambda args: math.atan(self._to_number(args[0])) if args else 0
        m["DEGREES"] = lambda args: math.degrees(self._to_number(args[0])) if args else 0
        m["RADIANS"] = lambda args: math.radians(self._to_number(args[0])) if args else 0
        m["FACT"] = lambda args: math.factorial(int(self._to_number(args[0]))) if args and 0 <= int(self._to_number(args[0])) <= 170 else "#NUM!"

        # Statistical
        m["AVERAGE"] = self._fn_average
        m["AVG"] = self._fn_average
        m["AVERAGEA"] = lambda args: sum(self._to_number(a) for a in args) / len(args) if args else 0
        m["COUNT"] = lambda args: sum(1 for a in args if isinstance(a, (int, float)) or (isinstance(a, str) and a.replace(".", "").replace("-", "").isdigit() and a.strip() != ""))
        m["COUNTA"] = lambda args: sum(1 for a in args if a is not None and a != "")
        m["MIN"] = lambda args: min([self._to_number(a) for a in args if a is not None and a != ""] or [0])
        m["MAX"] = lambda args: max([self._to_number(a) for a in args if a is not None and a != ""] or [0])
        m["MEDIAN"] = self._fn_median
        m["MODE"] = self._fn_mode
        m["MODE.SNGL"] = self._fn_mode
        m["STDEV"] = self._fn_stdev
        m["STDEV.S"] = self._fn_stdev
        m["STDEVP"] = self._fn_stdevp
        m["STDEV.P"] = self._fn_stdevp
        m["VAR"] = self._fn_var
        m["VAR.S"] = self._fn_var
        m["VARP"] = self._fn_varp
        m["VAR.P"] = self._fn_varp
        m["LARGE"] = lambda args: sorted([self._to_number(a) for a in args[:-1]], reverse=True)[int(self._to_number(args[-1])) - 1] if len(args) >= 2 and 1 <= int(self._to_number(args[-1])) <= len(args) - 1 else "#NUM!"
        m["SMALL"] = lambda args: sorted([self._to_number(a) for a in args[:-1]])[int(self._to_number(args[-1])) - 1] if len(args) >= 2 and 1 <= int(self._to_number(args[-1])) <= len(args) - 1 else "#NUM!"

        # Logical
        m["AND"] = lambda args: all(self._to_bool(a) for a in args) if args else True
        m["OR"] = lambda args: any(self._to_bool(a) for a in args) if args else False
        m["NOT"] = lambda args: not self._to_bool(args[0]) if args else True
        m["XOR"] = lambda args: sum(1 for a in args if self._to_bool(a)) % 2 == 1
        m["TRUE"] = lambda args: True
        m["FALSE"] = lambda args: False

        # Text
        m["CONCAT"] = lambda args: "".join(str(a if a is not None else "") for a in args)
        m["CONCATENATE"] = lambda args: "".join(str(a if a is not None else "") for a in args)
        m["TEXTJOIN"] = lambda args: str(args[0]).join([str(a) for a in args[2:] if not (self._to_bool(args[1]) and (a is None or a == ""))]) if len(args) >= 3 else ""
        m["LEFT"] = lambda args: str(args[0])[:int(self._to_number(args[1]))] if len(args) > 1 else str(args[0])[:1] if args else ""
        m["RIGHT"] = lambda args: str(args[0])[-int(self._to_number(args[1])):] if len(args) > 1 else str(args[0])[-1:] if args else ""
        m["MID"] = lambda args: str(args[0])[int(self._to_number(args[1])) - 1: int(self._to_number(args[1])) - 1 + int(self._to_number(args[2]))] if len(args) >= 3 else ""
        m["LEN"] = lambda args: len(str(args[0])) if args and args[0] is not None else 0
        m["TRIM"] = lambda args: " ".join(str(args[0]).split()) if args and args[0] is not None else ""
        m["UPPER"] = lambda args: str(args[0]).upper() if args and args[0] is not None else ""
        m["LOWER"] = lambda args: str(args[0]).lower() if args and args[0] is not None else ""
        m["PROPER"] = lambda args: str(args[0]).title() if args and args[0] is not None else ""
        m["EXACT"] = lambda args: str(args[0]) == str(args[1]) if len(args) >= 2 else False
        m["FIND"] = lambda args: str(args[1]).find(str(args[0]), int(self._to_number(args[2])) - 1 if len(args) > 2 else 0) + 1 if len(args) >= 2 and str(args[0]) in str(args[1]) else "#VALUE!"
        m["SEARCH"] = lambda args: str(args[1]).lower().find(str(args[0]).lower(), int(self._to_number(args[2])) - 1 if len(args) > 2 else 0) + 1 if len(args) >= 2 and str(args[0]).lower() in str(args[1]).lower() else "#VALUE!"
        m["REPLACE"] = self._fn_replace
        m["SUBSTITUTE"] = self._fn_substitute
        m["REPT"] = lambda args: str(args[0]) * int(self._to_number(args[1])) if len(args) >= 2 else ""
        m["TEXT"] = lambda args: str(args[0]) if args else ""
        m["VALUE"] = lambda args: self._to_number(args[0]) if args else 0
        m["CHAR"] = lambda args: chr(int(self._to_number(args[0]))) if args and 0 <= int(self._to_number(args[0])) <= 65535 else "#VALUE!"
        m["CODE"] = lambda args: ord(str(args[0])[0]) if args and str(args[0]) else "#VALUE!"
        m["CLEAN"] = lambda args: "".join(c for c in str(args[0]) if ord(c) >= 32) if args and args[0] else ""
        m["T"] = lambda args: str(args[0]) if args and isinstance(args[0], str) else ""

        # Date & Time
        m["TODAY"] = lambda args: date.today().isoformat()
        m["NOW"] = lambda args: datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        m["DATE"] = lambda args: date(int(self._to_number(args[0])), int(self._to_number(args[1])), int(self._to_number(args[2]))).isoformat() if len(args) >= 3 else "#VALUE!"
        m["YEAR"] = lambda args: self._parse_date(args[0]).year if args and self._parse_date(args[0]) else "#VALUE!"
        m["MONTH"] = lambda args: self._parse_date(args[0]).month if args and self._parse_date(args[0]) else "#VALUE!"
        m["DAY"] = lambda args: self._parse_date(args[0]).day if args and self._parse_date(args[0]) else "#VALUE!"
        m["DAYS"] = lambda args: (self._parse_date(args[0]) - self._parse_date(args[1])).days if len(args) >= 2 and self._parse_date(args[0]) and self._parse_date(args[1]) else "#VALUE!"

        # Information
        m["ISBLANK"] = lambda args: args[0] is None or args[0] == "" if args else True
        m["ISNUMBER"] = lambda args: isinstance(args[0], (int, float)) if args else False
        m["ISTEXT"] = lambda args: isinstance(args[0], str) and not args[0].replace(".", "").replace("-", "").isdigit() if args else False
        m["ISNONTEXT"] = lambda args: not (isinstance(args[0], str) and not args[0].replace(".", "").replace("-", "").isdigit()) if args else True
        m["ISLOGICAL"] = lambda args: isinstance(args[0], bool) if args else False
        m["ISERROR"] = lambda args: isinstance(args[0], str) and args[0].startswith("#") if args else False
        m["ISERR"] = lambda args: isinstance(args[0], str) and args[0].startswith("#") and args[0] != "#N/A" if args else False
        m["ISNA"] = lambda args: args[0] == "#N/A" if args else False
        m["TYPE"] = self._fn_type
        m["N"] = lambda args: self._to_number(args[0]) if args else 0

        # Custom multi-argument handlers
        m["IF"] = self._custom_if
        m["IFS"] = self._custom_ifs
        m["IFERROR"] = self._custom_iferror
        m["IFNA"] = self._custom_ifna
        m["SWITCH"] = self._custom_switch
        m["CHOOSE"] = self._custom_choose
        m["VLOOKUP"] = self._custom_vlookup
        m["HLOOKUP"] = self._custom_hlookup
        m["XLOOKUP"] = self._custom_xlookup
        m["INDEX"] = self._custom_index
        m["MATCH"] = self._custom_match
        m["COUNTIF"] = self._custom_countif
        m["COUNTIFS"] = self._custom_countifs
        m["SUMIF"] = self._custom_sumif
        m["SUMIFS"] = self._custom_sumifs
        m["AVERAGEIF"] = self._custom_averageif
        m["AVERAGEIFS"] = self._custom_averageifs
        m["MINIFS"] = self._custom_minifs
        m["MAXIFS"] = self._custom_maxifs
        m["COUNTBLANK"] = self._custom_countblank
        m["ROWS"] = self._custom_rows
        m["COLUMNS"] = self._custom_columns

        return m

    # -------------------------------------------------------------------------
    # Specialized Function Implementations
    # -------------------------------------------------------------------------

    def _fn_average(self, args: list) -> float:
        nums = [self._to_number(a) for a in args if a is not None and a != ""]
        return sum(nums) / len(nums) if nums else 0.0

    def _fn_median(self, args: list) -> float:
        nums = sorted(self._to_number(a) for a in args if a is not None and a != "")
        if not nums:
            return 0.0
        n = len(nums)
        mid = n // 2
        return (nums[mid - 1] + nums[mid]) / 2 if n % 2 == 0 else nums[mid]

    def _fn_mode(self, args: list) -> Any:
        nums = [self._to_number(a) for a in args if a is not None and a != ""]
        if not nums:
            return "#N/A"
        counts = {}
        for x in nums:
            counts[x] = counts.get(x, 0) + 1
        top = max(counts.values())
        if top <= 1:
            return "#N/A"
        for x in nums:
            if counts[x] == top:
                return x
        return "#N/A"

    def _fn_stdev(self, args: list) -> float:
        nums = [self._to_number(a) for a in args if a is not None and a != ""]
        if len(nums) < 2:
            return 0.0
        avg = sum(nums) / len(nums)
        var = sum((x - avg) ** 2 for x in nums) / (len(nums) - 1)
        return math.sqrt(var)

    def _fn_stdevp(self, args: list) -> float:
        nums = [self._to_number(a) for a in args if a is not None and a != ""]
        if not nums:
            return 0.0
        avg = sum(nums) / len(nums)
        var = sum((x - avg) ** 2 for x in nums) / len(nums)
        return math.sqrt(var)

    def _fn_var(self, args: list) -> float:
        nums = [self._to_number(a) for a in args if a is not None and a != ""]
        if len(nums) < 2:
            return 0.0
        avg = sum(nums) / len(nums)
        return sum((x - avg) ** 2 for x in nums) / (len(nums) - 1)

    def _fn_varp(self, args: list) -> float:
        nums = [self._to_number(a) for a in args if a is not None and a != ""]
        if not nums:
            return 0.0
        avg = sum(nums) / len(nums)
        return sum((x - avg) ** 2 for x in nums) / len(nums)

    def _fn_replace(self, args: list) -> str:
        if len(args) < 4:
            return "#VALUE!"
        old = str(args[0])
        start = max(1, int(self._to_number(args[1]))) - 1
        num = int(self._to_number(args[2]))
        new = str(args[3])
        return old[:start] + new + old[start + num:]

    def _fn_substitute(self, args: list) -> str:
        if len(args) < 3:
            return "#VALUE!"
        text, old_s, new_s = str(args[0]), str(args[1]), str(args[2])
        if len(args) >= 4:
            target_instance = int(self._to_number(args[3]))
            parts = text.split(old_s)
            if target_instance < 1 or target_instance >= len(parts):
                return text
            return old_s.join(parts[:target_instance]) + new_s + old_s.join(parts[target_instance:])
        return text.replace(old_s, new_s)

    def _fn_type(self, args: list) -> int:
        if not args or args[0] is None:
            return 1
        val = args[0]
        if isinstance(val, (int, float)): return 1
        if isinstance(val, str):
            return 16 if val.startswith("#") else 2
        if isinstance(val, bool): return 4
        if isinstance(val, list): return 64
        return 2

    def _parse_date(self, val: Any) -> date | None:
        if isinstance(val, (date, datetime)):
            return val.date() if isinstance(val, datetime) else val
        s = str(val).strip()
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        return None

    # -------------------------------------------------------------------------
    # Custom Control & Range Functions
    # -------------------------------------------------------------------------

    def _custom_if(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 2:
            return "#VALUE!"
        cond = engine._to_bool(engine._eval_expr(raw_args[0], current_sheet))
        if cond:
            return engine._eval_expr(raw_args[1], current_sheet)
        if len(raw_args) >= 3:
            return engine._eval_expr(raw_args[2], current_sheet)
        return False

    def _custom_ifs(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        for i in range(0, len(raw_args) - 1, 2):
            cond = engine._to_bool(engine._eval_expr(raw_args[i], current_sheet))
            if cond:
                return engine._eval_expr(raw_args[i + 1], current_sheet)
        return "#N/A"

    def _custom_iferror(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 2:
            return "#VALUE!"
        val = engine._eval_expr(raw_args[0], current_sheet)
        if isinstance(val, str) and val.startswith("#"):
            return engine._eval_expr(raw_args[1], current_sheet)
        return val

    def _custom_ifna(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 2:
            return "#VALUE!"
        val = engine._eval_expr(raw_args[0], current_sheet)
        if val == "#N/A":
            return engine._eval_expr(raw_args[1], current_sheet)
        return val

    def _custom_switch(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 3:
            return "#VALUE!"
        target = engine._eval_expr(raw_args[0], current_sheet)
        i = 1
        while i < len(raw_args) - 1:
            val = engine._eval_expr(raw_args[i], current_sheet)
            if engine._compare(target, val, "="):
                return engine._eval_expr(raw_args[i + 1], current_sheet)
            i += 2
        if i < len(raw_args):
            return engine._eval_expr(raw_args[i], current_sheet)
        return "#N/A"

    def _custom_choose(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 2:
            return "#VALUE!"
        idx = int(engine._to_number(engine._eval_expr(raw_args[0], current_sheet)))
        if 1 <= idx < len(raw_args):
            return engine._eval_expr(raw_args[idx], current_sheet)
        return "#VALUE!"

    def _get_2d_range(self, range_str: str, current_sheet: str | None) -> list[list[Any]]:
        """Returns 2D grid of values [row][col] for a range."""
        range_str = range_str.strip()
        match = RANGE_PATTERN.match(range_str)
        if not match:
            # Single cell
            if CELL_REF_PATTERN.match(range_str):
                cell = CellRef.from_string(range_str)
                sheet = cell.sheet or current_sheet
                return [[self._get_val(cell.row, cell.col, sheet)]]
            return [[]]

        sheet1, sheet2, start_str, end_str = match.groups()
        sheet = sheet1 or sheet2 or current_sheet
        start = CellRef.from_string(start_str)
        end = CellRef.from_string(end_str)

        min_r = min(start.row, end.row)
        max_r = max(start.row, end.row)
        min_c = min(start.col, end.col)
        max_c = max(start.col, end.col)

        grid = []
        for r in range(min_r, max_r + 1):
            row_vals = []
            for c in range(min_c, max_c + 1):
                row_vals.append(self._get_val(r, c, sheet))
            grid.append(row_vals)
        return grid

    def _custom_vlookup(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 3:
            return "#VALUE!"
        lookup_val = engine._eval_expr(raw_args[0], current_sheet)
        table = self._get_2d_range(raw_args[1], current_sheet)
        col_idx = int(engine._to_number(engine._eval_expr(raw_args[2], current_sheet))) - 1
        approx = engine._to_bool(engine._eval_expr(raw_args[3], current_sheet)) if len(raw_args) > 3 else False

        if not table or col_idx < 0:
            return "#REF!"

        for row in table:
            if not row:
                continue
            first_val = row[0]
            if engine._compare(first_val, lookup_val, "="):
                if col_idx < len(row):
                    return row[col_idx]
                return "#REF!"
        return "#N/A"

    def _custom_hlookup(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 3:
            return "#VALUE!"
        lookup_val = engine._eval_expr(raw_args[0], current_sheet)
        table = self._get_2d_range(raw_args[1], current_sheet)
        row_idx = int(engine._to_number(engine._eval_expr(raw_args[2], current_sheet))) - 1

        if not table or row_idx < 0 or row_idx >= len(table):
            return "#REF!"

        first_row = table[0]
        for c_idx, val in enumerate(first_row):
            if engine._compare(val, lookup_val, "="):
                if c_idx < len(table[row_idx]):
                    return table[row_idx][c_idx]
                return "#REF!"
        return "#N/A"

    def _custom_xlookup(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 3:
            return "#VALUE!"
        lookup_val = engine._eval_expr(raw_args[0], current_sheet)
        lookup_arr = [x for row in self._get_2d_range(raw_args[1], current_sheet) for x in row]
        return_arr = [x for row in self._get_2d_range(raw_args[2], current_sheet) for x in row]
        if_not_found = engine._eval_expr(raw_args[3], current_sheet) if len(raw_args) > 3 else "#N/A"

        for i, val in enumerate(lookup_arr):
            if engine._compare(val, lookup_val, "="):
                if i < len(return_arr):
                    return return_arr[i]
                return "#REF!"
        return if_not_found

    def _custom_index(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 2:
            return "#VALUE!"
        table = self._get_2d_range(raw_args[0], current_sheet)
        row_idx = int(engine._to_number(engine._eval_expr(raw_args[1], current_sheet))) - 1
        col_idx = int(engine._to_number(engine._eval_expr(raw_args[2], current_sheet))) - 1 if len(raw_args) >= 3 else 0

        if not table or row_idx < 0 or row_idx >= len(table):
            return "#REF!"
        if col_idx < 0 or col_idx >= len(table[row_idx]):
            return "#REF!"
        return table[row_idx][col_idx]

    def _custom_match(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 2:
            return "#VALUE!"
        lookup_val = engine._eval_expr(raw_args[0], current_sheet)
        table = self._get_2d_range(raw_args[1], current_sheet)
        items = [x for row in table for x in row]
        match_type = int(engine._to_number(engine._eval_expr(raw_args[2], current_sheet))) if len(raw_args) >= 3 else 1

        for i, val in enumerate(items):
            if engine._compare(val, lookup_val, "="):
                return i + 1
        return "#N/A"

    def _custom_countif(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> int:
        if len(raw_args) < 2:
            return 0
        vals = [x for row in self._get_2d_range(raw_args[0], current_sheet) for x in row]
        crit = engine._eval_expr(raw_args[1], current_sheet)
        return sum(1 for v in vals if self._match_criteria(v, crit))

    def _custom_countifs(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> int:
        if len(raw_args) < 2 or len(raw_args) % 2 != 0:
            return 0
        ranges = []
        criteria = []
        for i in range(0, len(raw_args), 2):
            ranges.append([x for row in self._get_2d_range(raw_args[i], current_sheet) for x in row])
            criteria.append(engine._eval_expr(raw_args[i + 1], current_sheet))

        n = min(len(r) for r in ranges) if ranges else 0
        count = 0
        for i in range(n):
            if all(self._match_criteria(ranges[k][i], criteria[k]) for k in range(len(ranges))):
                count += 1
        return count

    def _custom_sumif(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> float:
        if len(raw_args) < 2:
            return 0.0
        range_vals = [x for row in self._get_2d_range(raw_args[0], current_sheet) for x in row]
        crit = engine._eval_expr(raw_args[1], current_sheet)
        sum_vals = [x for row in self._get_2d_range(raw_args[2], current_sheet) for x in row] if len(raw_args) >= 3 else range_vals

        total = 0.0
        for i, val in enumerate(range_vals):
            if self._match_criteria(val, crit):
                s_val = sum_vals[i] if i < len(sum_vals) else 0
                total += self._to_number(s_val)
        return total

    def _custom_sumifs(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> float:
        if len(raw_args) < 3:
            return 0.0
        sum_vals = [x for row in self._get_2d_range(raw_args[0], current_sheet) for x in row]
        ranges = []
        criteria = []
        for i in range(1, len(raw_args) - 1, 2):
            ranges.append([x for row in self._get_2d_range(raw_args[i], current_sheet) for x in row])
            criteria.append(engine._eval_expr(raw_args[i + 1], current_sheet))

        n = min(len(sum_vals), *(len(r) for r in ranges)) if ranges else 0
        total = 0.0
        for i in range(n):
            if all(self._match_criteria(ranges[k][i], criteria[k]) for k in range(len(ranges))):
                total += self._to_number(sum_vals[i])
        return total

    def _custom_averageif(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 2:
            return 0.0
        range_vals = [x for row in self._get_2d_range(raw_args[0], current_sheet) for x in row]
        crit = engine._eval_expr(raw_args[1], current_sheet)
        avg_vals = [x for row in self._get_2d_range(raw_args[2], current_sheet) for x in row] if len(raw_args) >= 3 else range_vals

        nums = []
        for i, val in enumerate(range_vals):
            if self._match_criteria(val, crit):
                a_val = avg_vals[i] if i < len(avg_vals) else 0
                nums.append(self._to_number(a_val))
        return sum(nums) / len(nums) if nums else "#DIV/0!"

    def _custom_averageifs(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 3:
            return 0.0
        avg_vals = [x for row in self._get_2d_range(raw_args[0], current_sheet) for x in row]
        ranges = []
        criteria = []
        for i in range(1, len(raw_args) - 1, 2):
            ranges.append([x for row in self._get_2d_range(raw_args[i], current_sheet) for x in row])
            criteria.append(engine._eval_expr(raw_args[i + 1], current_sheet))

        n = min(len(avg_vals), *(len(r) for r in ranges)) if ranges else 0
        nums = []
        for i in range(n):
            if all(self._match_criteria(ranges[k][i], criteria[k]) for k in range(len(ranges))):
                nums.append(self._to_number(avg_vals[i]))
        return sum(nums) / len(nums) if nums else "#DIV/0!"

    def _custom_minifs(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 3:
            return 0.0
        min_vals = [x for row in self._get_2d_range(raw_args[0], current_sheet) for x in row]
        ranges = []
        criteria = []
        for i in range(1, len(raw_args) - 1, 2):
            ranges.append([x for row in self._get_2d_range(raw_args[i], current_sheet) for x in row])
            criteria.append(engine._eval_expr(raw_args[i + 1], current_sheet))

        n = min(len(min_vals), *(len(r) for r in ranges)) if ranges else 0
        nums = []
        for i in range(n):
            if all(self._match_criteria(ranges[k][i], criteria[k]) for k in range(len(ranges))):
                nums.append(self._to_number(min_vals[i]))
        return min(nums) if nums else 0.0

    def _custom_maxifs(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> Any:
        if len(raw_args) < 3:
            return 0.0
        max_vals = [x for row in self._get_2d_range(raw_args[0], current_sheet) for x in row]
        ranges = []
        criteria = []
        for i in range(1, len(raw_args) - 1, 2):
            ranges.append([x for row in self._get_2d_range(raw_args[i], current_sheet) for x in row])
            criteria.append(engine._eval_expr(raw_args[i + 1], current_sheet))

        n = min(len(max_vals), *(len(r) for r in ranges)) if ranges else 0
        nums = []
        for i in range(n):
            if all(self._match_criteria(ranges[k][i], criteria[k]) for k in range(len(ranges))):
                nums.append(self._to_number(max_vals[i]))
        return max(nums) if nums else 0.0

    def _custom_countblank(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> int:
        if not raw_args:
            return 0
        vals = [x for row in self._get_2d_range(raw_args[0], current_sheet) for x in row]
        return sum(1 for v in vals if v is None or v == "")

    def _custom_rows(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> int:
        if not raw_args:
            return 0
        table = self._get_2d_range(raw_args[0], current_sheet)
        return len(table)

    def _custom_columns(self, raw_args: list[str], current_sheet: str | None, engine: "FormulaEngine") -> int:
        if not raw_args:
            return 0
        table = self._get_2d_range(raw_args[0], current_sheet)
        return len(table[0]) if table else 0


# =============================================================================
# Formula Store
# =============================================================================

class FormulaStore:
    """Stores cell formulas per sheet."""

    def __init__(self) -> None:
        # sheet_name -> {(row, col): formula_string}
        self._formulas: dict[str, dict[tuple[int, int], str]] = {}

    def set(self, sheet: str, row: int, col: int, formula: str) -> None:
        if sheet not in self._formulas:
            self._formulas[sheet] = {}
        if formula and str(formula).startswith("="):
            self._formulas[sheet][(row, col)] = formula
        elif (row, col) in self._formulas[sheet]:
            del self._formulas[sheet][(row, col)]

    def get(self, sheet: str, row: int, col: int) -> str | None:
        return self._formulas.get(sheet, {}).get((row, col))

    def has_formula(self, sheet: str, row: int, col: int) -> bool:
        return (row, col) in self._formulas.get(sheet, {})

    def get_all(self, sheet: str) -> dict[tuple[int, int], str]:
        return self._formulas.get(sheet, {}).copy()

    def clear_sheet(self, sheet: str) -> None:
        if sheet in self._formulas:
            self._formulas[sheet].clear()

    def clear_all(self) -> None:
        self._formulas.clear()

    def rename_sheet(self, old_name: str, new_name: str) -> None:
        if old_name in self._formulas:
            self._formulas[new_name] = self._formulas.pop(old_name)

    def delete_row(self, sheet: str, row_idx: int) -> None:
        if sheet not in self._formulas:
            return
        to_del = [k for k in self._formulas[sheet] if k[0] == row_idx]
        for k in to_del:
            del self._formulas[sheet][k]
        new_dict = {}
        for (r, c), f in self._formulas[sheet].items():
            if r > row_idx:
                new_dict[(r - 1, c)] = shift_formula_references(f, -1, 0)
            else:
                new_dict[(r, c)] = f
        self._formulas[sheet] = new_dict

    def insert_row(self, sheet: str, row_idx: int) -> None:
        if sheet not in self._formulas:
            return
        new_dict = {}
        for (r, c), f in self._formulas[sheet].items():
            if r >= row_idx:
                new_dict[(r + 1, c)] = shift_formula_references(f, 1, 0)
            else:
                new_dict[(r, c)] = f
        self._formulas[sheet] = new_dict

    def delete_col(self, sheet: str, col_idx: int) -> None:
        if sheet not in self._formulas:
            return
        to_del = [k for k in self._formulas[sheet] if k[1] == col_idx]
        for k in to_del:
            del self._formulas[sheet][k]
        new_dict = {}
        for (r, c), f in self._formulas[sheet].items():
            if c > col_idx:
                new_dict[(r, c - 1)] = shift_formula_references(f, 0, -1)
            else:
                new_dict[(r, c)] = f
        self._formulas[sheet] = new_dict

    def insert_col(self, sheet: str, col_idx: int) -> None:
        if sheet not in self._formulas:
            return
        new_dict = {}
        for (r, c), f in self._formulas[sheet].items():
            if c >= col_idx:
                new_dict[(r, c + 1)] = shift_formula_references(f, 0, 1)
            else:
                new_dict[(r, c)] = f
        self._formulas[sheet] = new_dict
