"""
Data models and state classes for Excel Viewer Pro.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Any, Iterator
import re
from openpyxl.utils import get_column_letter, column_index_from_string


# =============================================================================
# Cell Coordinates and Ranges
# =============================================================================

CELL_REF_REGEX = re.compile(r"^\$?([A-Za-z]+)\$?(\d+)$")
RANGE_REF_REGEX = re.compile(r"^(\$?[A-Za-z]+\$?\d+):(\$?[A-Za-z]+\$?\d+)$")


@dataclass(slots=True)
class CellPosition:
    """Represents a 0-based cell coordinate (row, col)."""
    row: int = 0
    col: int = 0

    def to_excel(self) -> str:
        """Convert to Excel notation (e.g. A1, B2)."""
        return f"{get_column_letter(self.col + 1)}{self.row + 1}"

    @classmethod
    def from_excel(cls, ref: str) -> "CellPosition":
        """Parse Excel notation (e.g. A1, $B$2) to CellPosition."""
        clean = ref.replace("$", "").strip()
        match = CELL_REF_REGEX.match(clean)
        if not match:
            raise ValueError(f"Invalid cell reference: {ref}")
        col_str, row_str = match.groups()
        return cls(
            row=int(row_str) - 1,
            col=column_index_from_string(col_str.upper()) - 1
        )

    def copy(self) -> "CellPosition":
        return CellPosition(self.row, self.col)

    def __iter__(self):
        yield self.row
        yield self.col


@dataclass(slots=True)
class CellRange:
    """Represents a rectangular range of cells (anchor and active/extent)."""
    start: CellPosition = field(default_factory=lambda: CellPosition(0, 0))
    end: CellPosition = field(default_factory=lambda: CellPosition(0, 0))

    def normalized(self) -> tuple[int, int, int, int]:
        """Returns (min_row, max_row, min_col, max_col)."""
        min_r = min(self.start.row, self.end.row)
        max_r = max(self.start.row, self.end.row)
        min_c = min(self.start.col, self.end.col)
        max_c = max(self.start.col, self.end.col)
        return min_r, max_r, min_c, max_c

    def to_excel(self) -> str:
        """Return range string like 'A1:C10' or 'A1' if single cell."""
        min_r, max_r, min_c, max_c = self.normalized()
        p1 = CellPosition(min_r, min_c).to_excel()
        if min_r == max_r and min_c == max_c:
            return p1
        p2 = CellPosition(max_r, max_c).to_excel()
        return f"{p1}:{p2}"

    @classmethod
    def from_excel(cls, range_str: str) -> "CellRange":
        """Parse 'A1:B10' or 'A1' into CellRange."""
        clean = range_str.strip()
        if ":" in clean:
            s1, s2 = clean.split(":", 1)
            return cls(CellPosition.from_excel(s1), CellPosition.from_excel(s2))
        pos = CellPosition.from_excel(clean)
        return cls(pos.copy(), pos.copy())

    def contains(self, row: int, col: int) -> bool:
        """Check if row and col are inside the range."""
        min_r, max_r, min_c, max_c = self.normalized()
        return min_r <= row <= max_r and min_c <= col <= max_c

    def iter_cells(self) -> Iterator[tuple[int, int]]:
        """Yield (row, col) for all cells in range in row-major order."""
        min_r, max_r, min_c, max_c = self.normalized()
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                yield r, c

    @property
    def row_count(self) -> int:
        min_r, max_r, _, _ = self.normalized()
        return max_r - min_r + 1

    @property
    def col_count(self) -> int:
        _, _, min_c, max_c = self.normalized()
        return max_c - min_c + 1

    @property
    def is_single_cell(self) -> bool:
        min_r, max_r, min_c, max_c = self.normalized()
        return min_r == max_r and min_c == max_c


# =============================================================================
# Cell Styling & Formatting Model
# =============================================================================

@dataclass
class CellStyle:
    """Styling properties for a cell."""
    font_name: str = "Calibri"
    font_size: int = 11
    bold: bool = False
    italic: bool = False
    underline: bool = False
    strikethrough: bool = False
    fg_color: str | None = None  # Hex '#000000'
    bg_color: str | None = None  # Hex '#FFFFFF'
    halign: str = "left"  # left, center, right
    valign: str = "center"  # top, center, bottom
    wrap_text: bool = False
    number_format: str = ""
    borders: dict[str, str] = field(default_factory=dict)  # 'top': 'thin', 'bottom': 'double', etc.

    def copy(self) -> "CellStyle":
        return CellStyle(
            font_name=self.font_name,
            font_size=self.font_size,
            bold=self.bold,
            italic=self.italic,
            underline=self.underline,
            strikethrough=self.strikethrough,
            fg_color=self.fg_color,
            bg_color=self.bg_color,
            halign=self.halign,
            valign=self.valign,
            wrap_text=self.wrap_text,
            number_format=self.number_format,
            borders=dict(self.borders)
        )


@dataclass
class CellComment:
    """Comment / Note attached to a cell."""
    text: str = ""
    author: str = ""
    timestamp: str = ""


# =============================================================================
# Undo / Redo Actions
# =============================================================================

@dataclass
class UndoAction:
    """Action recorded for undo/redo."""
    action_type: str  # 'cell_change', 'range_change', 'insert_row', 'delete_row', 'insert_col', 'delete_col', 'format_change'
    sheet_name: str
    data: dict[str, Any] = field(default_factory=dict)


# =============================================================================
# Sheet Data Model
# =============================================================================

@dataclass
class SheetData:
    """Complete in-memory model for a worksheet."""
    name: str = "Sheet1"
    headers: list[str] = field(default_factory=list)
    rows: list[list[Any]] = field(default_factory=list)
    col_count: int = 0
    row_count: int = 0
    cell_styles: dict[tuple[int, int], CellStyle] = field(default_factory=dict)
    comments: dict[tuple[int, int], CellComment] = field(default_factory=dict)
    hidden_rows: set[int] = field(default_factory=set)
    hidden_cols: set[int] = field(default_factory=set)
    column_widths: dict[int, int] = field(default_factory=dict)
    row_heights: dict[int, int] = field(default_factory=dict)
    merged_ranges: list[CellRange] = field(default_factory=list)
    filter_active: bool = False
    filter_criteria: dict[int, set[str]] = field(default_factory=dict)

    def clear(self) -> None:
        self.headers.clear()
        self.rows.clear()
        self.col_count = 0
        self.row_count = 0
        self.cell_styles.clear()
        self.comments.clear()
        self.hidden_rows.clear()
        self.hidden_cols.clear()
        self.column_widths.clear()
        self.row_heights.clear()
        self.merged_ranges.clear()
        self.filter_active = False
        self.filter_criteria.clear()

    def get_cell_value(self, row: int, col: int) -> Any:
        """Safely get cell value."""
        if 0 <= row < len(self.rows) and 0 <= col < len(self.rows[row]):
            return self.rows[row][col]
        return None

    def set_cell_value(self, row: int, col: int, value: Any) -> None:
        """Safely set cell value expanding rows/cols if necessary."""
        while row >= len(self.rows):
            self.rows.append(["" for _ in range(self.col_count)])
        
        while col >= len(self.rows[row]):
            self.rows[row].append("")
            
        self.rows[row][col] = value
        self.row_count = len(self.rows)
        if col + 1 > self.col_count:
            self.col_count = col + 1
            while len(self.headers) < self.col_count:
                self.headers.append(get_column_letter(len(self.headers) + 1))
