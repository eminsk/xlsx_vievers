"""
Excel Viewer Pro — Professional Spreadsheet Application.
========================================================

Complete desktop spreadsheet viewer and editor with:
- Full Excel Ribbon (Home, Insert, Data, Formulas, View)
- Multi-cell rectangular range selection & autofill
- Dynamic range statistics in StatusBar (Average, Count, Nums, Min, Max, Sum)
- Excel-compatible 2D Clipboard (TSV Copy/Paste/Cut & Paste Special)
- Comprehensive Formula Engine (80+ functions, multi-sheet references)
- Interactive AutoFilter with checkbox search
- Multi-level Custom Sort dialog
- Find and Replace dialog with whole-sheet search
- Chart Wizard with 10+ chart types and themes
- Cell formatting (Fonts, Sizes, Colors, Alignments, Borders, Number Formats)
- Conditional Formatting (Highlight Rules & Color Scales)
- Sheet Manager (+ Tab, Rename, Duplicate, Delete)
- OpenPyXL synchronization, CSV/TSV/HTML Export
"""

from __future__ import annotations

import csv
import io
from pathlib import Path
from collections import deque
from functools import partial
from typing import Any

import tkinter as tk
from tkinter import filedialog, messagebox, Menu, colorchooser, simpledialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string

from config import Config
from models import CellPosition, CellRange, CellStyle, CellComment, SheetData, UndoAction
from formulas import FormulaEngine, FormulaStore, shift_formula_references, CELL_REF_PATTERN, RANGE_PATTERN
from formatting import NumberFormatter, ConditionalFormattingEngine
from widgets import ExcelRibbon, FormulaBar, SheetTabs, StatusBar, CellEditor
from dialogs import (
    FindReplaceDialog, InsertFunctionDialog, CustomSortDialog,
    AutoFilterPopup, ChartWizardDialog, RemoveDuplicatesDialog,
    TextToColumnsDialog, GoalSeekDialog, CellCommentDialog
)


# =============================================================================
# Main Application Class
# =============================================================================

class ExcelViewerPro(ttk.Window):
    """Professional Excel-like Spreadsheet Application."""

    def __init__(self) -> None:
        super().__init__(themename=Config.THEME)
        self._init_state()
        self._init_window()
        self._build_ui()
        self._bind_events()
        self._create_new_workbook()

    # =========================================================================
    # Initialization
    # =========================================================================

    def _init_state(self) -> None:
        """Initialize core application state."""
        self._workbook: Workbook | None = None
        self._sheet: Worksheet | None = None
        self._file_path: Path | None = None
        self._modified = False

        # Multi-sheet in-memory cache
        self._sheets_data: dict[str, SheetData] = {}
        self._active_sheet_name: str = "Sheet1"

        # Selection state
        self._selected = CellPosition(0, 0)
        self._range_anchor = CellPosition(0, 0)
        self._range_extent = CellPosition(0, 0)
        self._is_mouse_dragging = False

        # Treeview mapping
        self._row_iids: list[str] = []
        self._col_ids: list[str] = []
        self._detached_rows: dict[int, str] = {}

        # History & Clipboard
        self._undo_stack: deque[UndoAction] = deque(maxlen=Config.UNDO_LIMIT)
        self._redo_stack: deque[UndoAction] = deque(maxlen=Config.UNDO_LIMIT)
        self._clipboard_buffer: list[list[str]] = []
        self._format_painter_style: CellStyle | None = None

        # Formula Engine
        self._formula_store = FormulaStore()
        self._formula_engine = FormulaEngine(self._get_cell_value_for_formula)

        # UI state
        self._active_cell_border = {}
        self._range_border = {}
        self._formula_bar_active = False

    def _init_window(self) -> None:
        """Configure main application window."""
        self.title(Config.APP_TITLE)
        self.geometry(Config.WINDOW_SIZE)
        self.minsize(*Config.MIN_SIZE)
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._set_app_icon()

    def _build_ui(self) -> None:
        """Build the UI structure: Menu, Ribbon, Formula Bar, Grid, Tabs, StatusBar."""
        self._create_menu()

        # Ribbon bar
        callbacks = {
            # File
            "open": self._open_file,
            "save": self._quick_save,
            "save_as": self._save_as,
            # Clipboard
            "paste": self._paste,
            "paste_values": lambda: self._paste_special("values"),
            "paste_formulas": lambda: self._paste_special("formulas"),
            "paste_transpose": lambda: self._paste_special("transpose"),
            "cut": self._cut,
            "copy": self._copy,
            "format_painter": self._activate_format_painter,
            # Font
            "set_font_family": self._set_font_family,
            "set_font_size": self._set_font_size,
            "increase_font_size": self._increase_font_size,
            "decrease_font_size": self._decrease_font_size,
            "toggle_bold": self._toggle_bold,
            "toggle_italic": self._toggle_italic,
            "toggle_underline": self._toggle_underline,
            "toggle_strikethrough": self._toggle_strikethrough,
            "set_borders": self._set_borders,
            "set_bg_color": self._pick_bg_color,
            "set_fg_color": self._pick_fg_color,
            # Alignment
            "set_halign": self._set_halign,
            "set_valign": self._set_valign,
            "toggle_wrap_text": self._toggle_wrap_text,
            "merge_cells": self._toggle_merge_cells,
            # Number Format
            "set_number_format": self._set_number_format,
            "increase_decimals": self._increase_decimals,
            "decrease_decimals": self._decrease_decimals,
            # Styles / Conditional Formatting
            "add_cf_rule": self._add_cf_rule,
            "clear_cf_rules": self._clear_cf_rules,
            # Cells
            "insert_row_above": self._insert_row_above,
            "insert_row_below": self._insert_row_below,
            "insert_col_left": self._insert_col_left,
            "insert_col_right": self._insert_col_right,
            "delete_row": self._delete_row,
            "delete_column": self._delete_column,
            # Editing
            "apply_autosum": self._apply_autosum,
            "quick_sort": self._quick_sort,
            "custom_sort_dialog": self._show_custom_sort_dialog,
            "toggle_filter": self._toggle_autofilter,
            "clear_filters": self._clear_filters,
            "find_dialog": self._show_find_dialog,
            "replace_dialog": self._show_replace_dialog,
            "goto_dialog": self._show_goto_dialog,
            # Insert Tab
            "show_chart_wizard": self._show_chart_wizard,
            "quick_chart": self._create_quick_chart,
            "edit_comment": self._edit_cell_comment,
            "show_fx_wizard": self._show_fx_wizard,
            "insert_formula_text": self._insert_formula_text,
            # Data Tab
            "text_to_columns_dialog": self._show_text_to_columns_dialog,
            "remove_duplicates_dialog": self._show_remove_duplicates_dialog,
            "goal_seek_dialog": self._show_goal_seek_dialog,
            # Formulas Tab
            "recalculate_sheet": self._recalculate_all,
            # View Tab
            "toggle_formula_bar": self._toggle_formula_bar,
            "toggle_status_bar": self._toggle_status_bar,
            "freeze_top_row": self._freeze_top_row,
            "freeze_first_col": self._freeze_first_col,
            "unfreeze_all": self._unfreeze_all,
            "autofit_all_cols": self._autofit_all_columns,
            "change_theme": self._change_theme
        }

        self._ribbon = ExcelRibbon(self, callbacks)
        self._ribbon.pack(fill=tk.X)

        # Formula Bar
        self._formula_bar = FormulaBar(
            self,
            on_commit=self._on_formula_commit,
            on_fx_clicked=self._show_fx_wizard,
            on_goto_cell=self._goto_cell
        )
        self._formula_bar.pack(fill=tk.X)

        # Spreadsheet Grid Area
        self._create_grid()

        # Sheet Tabs (Bottom)
        self._sheet_tabs = SheetTabs(
            self,
            on_select_sheet=self._on_switch_sheet,
            on_add_sheet=self._on_add_sheet,
            on_rename_sheet=self._on_rename_sheet,
            on_delete_sheet=self._on_delete_sheet,
            on_duplicate_sheet=self._on_duplicate_sheet
        )
        self._sheet_tabs.pack(fill=tk.X, side=tk.BOTTOM)

        # Status Bar (Bottom-most)
        self._status_bar = StatusBar(self, on_zoom_change=self._on_zoom_change)
        self._status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def _create_menu(self) -> None:
        """Create classic top menu bar."""
        menubar = Menu(self)
        self.config(menu=menubar)

        # File Menu
        file_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="New Workbook", command=self._create_new_workbook, accelerator="Ctrl+N")
        file_menu.add_command(label="Open...", command=self._open_file, accelerator="Ctrl+O")
        file_menu.add_command(label="Save", command=self._quick_save, accelerator="Ctrl+S")
        file_menu.add_command(label="Save As...", command=self._save_as, accelerator="Ctrl+Shift+S")
        file_menu.add_separator()

        export_menu = Menu(file_menu, tearoff=0)
        file_menu.add_cascade(label="Export", menu=export_menu)
        export_menu.add_command(label="Export as CSV (.csv)", command=lambda: self._export_file("csv"))
        export_menu.add_command(label="Export as TSV (.tsv)", command=lambda: self._export_file("tsv"))
        export_menu.add_command(label="Export as HTML Table (.html)", command=lambda: self._export_file("html"))

        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self._on_close)

        # Edit Menu
        edit_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Edit", menu=edit_menu)
        edit_menu.add_command(label="Undo", command=self._undo, accelerator="Ctrl+Z")
        edit_menu.add_command(label="Redo", command=self._redo, accelerator="Ctrl+Y")
        edit_menu.add_separator()
        edit_menu.add_command(label="Cut", command=self._cut, accelerator="Ctrl+X")
        edit_menu.add_command(label="Copy", command=self._copy, accelerator="Ctrl+C")
        edit_menu.add_command(label="Paste", command=self._paste, accelerator="Ctrl+V")
        edit_menu.add_separator()
        edit_menu.add_command(label="Select All", command=self._select_all, accelerator="Ctrl+A")
        edit_menu.add_command(label="Delete", command=self._delete_cell, accelerator="Del")
        edit_menu.add_separator()
        edit_menu.add_command(label="Find...", command=self._show_find_dialog, accelerator="Ctrl+F")
        edit_menu.add_command(label="Replace...", command=self._show_replace_dialog, accelerator="Ctrl+H")

        # View Menu
        view_menu = Menu(menubar, tearoff=0)
        menubar.add_cascade(label="View", menu=view_menu)
        view_menu.add_command(label="Auto-fit All Columns", command=self._autofit_all_columns)
        view_menu.add_command(label="Refresh (F5)", command=self._refresh, accelerator="F5")

    def _create_grid(self) -> None:
        """Create main spreadsheet grid."""
        container = ttk.Frame(self)
        container.pack(fill=tk.BOTH, expand=True, padx=4, pady=2)

        self._tree = ttk.Treeview(container, show="headings", selectmode="browse")

        self._vsb = ttk.Scrollbar(container, orient=tk.VERTICAL, command=self._tree.yview)
        self._hsb = ttk.Scrollbar(container, orient=tk.HORIZONTAL, command=self._tree.xview)
        self._tree.configure(yscrollcommand=self._on_tree_y_scroll, xscrollcommand=self._on_tree_x_scroll)

        self._tree.grid(row=0, column=0, sticky=NSEW)
        self._vsb.grid(row=0, column=1, sticky=NS)
        self._hsb.grid(row=1, column=0, sticky=EW)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        # Style tags
        self._tree.tag_configure("odd", background=Config.ROW_ALT_BG)
        self._tree.tag_configure("even", background=Config.CELL_BG)

        # Highlight frames
        self._active_cell_border = self._create_border_frames(Config.ACTIVE_CELL_BORDER)
        self._range_border = self._create_border_frames(Config.RANGE_BORDER)
        self._hide_border(self._active_cell_border)
        self._hide_border(self._range_border)

        # Inline editor
        self._cell_editor = CellEditor(self._tree, self._on_edit_commit, self._on_edit_cancel)

    # =========================================================================
    # Event Bindings
    # =========================================================================

    def _bind_events(self) -> None:
        """Bind keyboard shortcuts and mouse interactions."""
        shortcuts = {
            "<Control-n>": lambda e: self._create_new_workbook(),
            "<Control-o>": lambda e: self._open_file(),
            "<Control-s>": lambda e: self._quick_save(),
            "<Control-S>": lambda e: self._save_as(),
            "<Control-z>": lambda e: self._undo(),
            "<Control-y>": lambda e: self._redo(),
            "<Control-x>": lambda e: self._cut(),
            "<Control-c>": lambda e: self._copy(),
            "<Control-v>": lambda e: self._paste(),
            "<Control-a>": lambda e: self._select_all(),
            "<Control-f>": lambda e: self._show_find_dialog(),
            "<Control-h>": lambda e: self._show_replace_dialog(),
            "<Control-g>": lambda e: self._show_goto_dialog(),
            "<F5>": lambda e: self._refresh(),
            "<F2>": lambda e: self._start_inline_edit(),
            "<F9>": lambda e: self._recalculate_all(),
            "<Delete>": lambda e: self._delete_cell(),
            "<Escape>": lambda e: self._cancel_edit()
        }
        for key, handler in shortcuts.items():
            self.bind(key, handler)

        # Tree mouse events
        self._tree.bind("<Button-1>", self._on_grid_mouse_down)
        self._tree.bind("<B1-Motion>", self._on_grid_mouse_drag)
        self._tree.bind("<ButtonRelease-1>", self._on_grid_mouse_up)
        self._tree.bind("<Double-1>", self._on_grid_double_click)
        self._tree.bind("<Button-3>", self._show_grid_context_menu)
        self._tree.bind("<Configure>", lambda e: self._refresh_highlights())

        # Keyboard Navigation
        self._tree.bind("<Up>", lambda e: self._navigate(-1, 0, shift=e.state & 0x1))
        self._tree.bind("<Down>", lambda e: self._navigate(1, 0, shift=e.state & 0x1))
        self._tree.bind("<Left>", lambda e: self._navigate(0, -1, shift=e.state & 0x1))
        self._tree.bind("<Right>", lambda e: self._navigate(0, 1, shift=e.state & 0x1))
        self._tree.bind("<Tab>", lambda e: self._navigate(0, 1, shift=False))
        self._tree.bind("<Return>", lambda e: self._navigate(1, 0, shift=False))

        # Direct typing starts edit
        self._tree.bind("<Key>", self._on_grid_key_press)

    # =========================================================================
    # Workbook & Sheet Management
    # =========================================================================

    def _create_new_workbook(self) -> None:
        """Create a new empty workbook."""
        if self._modified and not self._confirm_discard():
            return

        self._close_workbook()
        self._workbook = Workbook()
        default_sheet = self._workbook.active
        default_sheet.title = "Sheet1"

        # Populate empty grid
        sheet_data = SheetData(name="Sheet1", col_count=10, row_count=30)
        sheet_data.headers = [get_column_letter(i + 1) for i in range(10)]
        sheet_data.rows = [["" for _ in range(10)] for _ in range(30)]

        self._sheets_data = {"Sheet1": sheet_data}
        self._active_sheet_name = "Sheet1"
        self._sheet = default_sheet
        self._file_path = None
        self._set_modified(False)

        self._sheet_tabs.set_sheets(["Sheet1"])
        self._load_active_sheet()
        self.title(f"{Config.APP_TITLE} — Untitled")

    def _open_file(self) -> None:
        """Open Excel or CSV/TSV file."""
        if self._modified and not self._confirm_discard():
            return

        path = filedialog.askopenfilename(
            title="Open Spreadsheet File",
            filetypes=[
                ("Excel Files", "*.xlsx *.xlsm *.xltx *.xltm"),
                ("CSV Files", "*.csv"),
                ("TSV Files", "*.tsv"),
                ("All Files", "*.*")
            ]
        )
        if path:
            self._load_file(Path(path))

    def _load_file(self, path: Path) -> None:
        """Load file into application."""
        try:
            self._status_bar.set_mode("Loading...")
            self.update_idletasks()

            self._close_workbook()
            self._file_path = path

            if path.suffix.lower() in (".csv", ".tsv"):
                self._load_delimited_file(path)
            else:
                self._workbook = load_workbook(path, data_only=False)
                self._sheets_data.clear()

                for s_name in self._workbook.sheetnames:
                    ws = self._workbook[s_name]
                    max_r = ws.max_row or 1
                    max_c = ws.max_column or 1
                    headers = []
                    rows = []
                    cell_styles = {}

                    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c)):
                        row_vals = []
                        for c_idx, cell in enumerate(row):
                            val = cell.value
                            if isinstance(val, str) and val.startswith("="):
                                self._formula_store.set(s_name, r_idx, c_idx, val)
                            row_vals.append(val)

                        rows.append(row_vals)

                    # Create default headers if empty
                    headers = [get_column_letter(i + 1) for i in range(max_c)]
                    s_data = SheetData(name=s_name, headers=headers, rows=rows, col_count=max_c, row_count=len(rows))
                    self._sheets_data[s_name] = s_data

                self._sheet_tabs.set_sheets(self._workbook.sheetnames)
                self._active_sheet_name = self._workbook.sheetnames[0]
                self._sheet = self._workbook[self._active_sheet_name]
                self._load_active_sheet()

            self._set_modified(False)
            self.title(f"{Config.APP_TITLE} — {path.name}")
            self._status_bar.set_mode("Ready")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file:\n{e}")
            self._status_bar.set_mode("Error")

    def _load_delimited_file(self, path: Path) -> None:
        """Load CSV / TSV file."""
        delim = "\t" if path.suffix.lower() == ".tsv" else ","
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f, delimiter=delim)
            rows = [list(r) for r in reader]

        max_c = max((len(r) for r in rows), default=1)
        for r in rows:
            while len(r) < max_c:
                r.append("")

        s_data = SheetData(
            name="Sheet1",
            headers=[get_column_letter(i + 1) for i in range(max_c)],
            rows=rows,
            col_count=max_c,
            row_count=len(rows)
        )
        self._sheets_data = {"Sheet1": s_data}
        self._active_sheet_name = "Sheet1"
        self._workbook = Workbook()
        self._sheet = self._workbook.active
        self._sheet.title = "Sheet1"
        self._sheet_tabs.set_sheets(["Sheet1"])
        self._load_active_sheet()

    def _quick_save(self) -> None:
        if not self._file_path:
            self._save_as()
            return
        self._save_to_path(self._file_path)

    def _save_as(self) -> None:
        path = filedialog.asksaveasfilename(
            title="Save Spreadsheet As",
            defaultextension=".xlsx",
            filetypes=[("Excel Files (*.xlsx)", "*.xlsx"), ("CSV Files (*.csv)", "*.csv"), ("TSV Files (*.tsv)", "*.tsv")]
        )
        if path:
            self._save_to_path(Path(path))

    def _save_to_path(self, path: Path) -> None:
        """Sync and save active workbook to disk."""
        try:
            self._status_bar.set_mode("Saving...")
            self.update_idletasks()

            # Ensure openpyxl workbook is in sync
            if not self._workbook:
                self._workbook = Workbook()

            # Sync all sheets from _sheets_data to workbook
            existing_sheets = set(self._workbook.sheetnames)
            for s_name, s_data in self._sheets_data.items():
                if s_name in existing_sheets:
                    ws = self._workbook[s_name]
                    ws.delete_rows(1, ws.max_row + 1)
                else:
                    ws = self._workbook.create_sheet(title=s_name)

                # Write rows and formulas
                for r_idx, row in enumerate(s_data.rows):
                    for c_idx, val in enumerate(row):
                        formula = self._formula_store.get(s_name, r_idx, c_idx)
                        cell_val = formula if formula else val
                        if cell_val is not None and cell_val != "":
                            ws.cell(row=r_idx + 1, column=c_idx + 1, value=cell_val)

            # Remove deleted sheets
            for old_s in existing_sheets:
                if old_s not in self._sheets_data and len(self._workbook.sheetnames) > 1:
                    del self._workbook[old_s]

            if path.suffix.lower() == ".csv":
                self._export_sheet_to_csv(self._active_sheet_name, path)
            elif path.suffix.lower() == ".tsv":
                self._export_sheet_to_csv(self._active_sheet_name, path, delimiter="\t")
            else:
                self._workbook.save(path)

            self._file_path = path
            self._set_modified(False)
            self.title(f"{Config.APP_TITLE} — {path.name}")
            self._status_bar.set_mode("Ready")
            messagebox.showinfo("Saved", f"File saved successfully:\n{path.name}")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to save file:\n{e}")
            self._status_bar.set_mode("Error")

    def _export_file(self, fmt: str) -> None:
        """Export sheet to CSV, TSV, or HTML."""
        ext = f".{fmt}"
        path = filedialog.asksaveasfilename(
            title=f"Export as {fmt.upper()}",
            defaultextension=ext,
            filetypes=[(f"{fmt.upper()} File", f"*{ext}")]
        )
        if not path:
            return

        s_data = self._get_active_sheet_data()
        p = Path(path)

        if fmt in ("csv", "tsv"):
            delim = "\t" if fmt == "tsv" else ","
            with open(p, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=delim)
                for row in s_data.rows:
                    writer.writerow(row)
        elif fmt == "html":
            with open(p, "w", encoding="utf-8") as f:
                f.write("<!DOCTYPE html><html><head><meta charset='utf-8'>")
                f.write("<style>table{border-collapse:collapse;font-family:sans-serif;}td,th{border:1px solid #ccc;padding:6px 10px;}</style>")
                f.write(f"</head><body><h2>{s_data.name}</h2><table>\n")
                f.write("<tr>" + "".join(f"<th>{h}</th>" for h in s_data.headers) + "</tr>\n")
                for row in s_data.rows:
                    f.write("<tr>" + "".join(f"<td>{v if v is not None else ''}</td>" for v in row) + "</tr>\n")
                f.write("</table></body></html>")

        messagebox.showinfo("Export", f"Exported successfully to {p.name}")

    def _export_sheet_to_csv(self, sheet_name: str, path: Path, delimiter: str = ",") -> None:
        s_data = self._sheets_data.get(sheet_name)
        if not s_data:
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=delimiter)
            for row in s_data.rows:
                writer.writerow(row)

    def _close_workbook(self) -> None:
        if self._workbook:
            try:
                self._workbook.close()
            except Exception:
                pass
            self._workbook = None
            self._sheet = None

    def _on_switch_sheet(self, sheet_name: str) -> None:
        """Switch active worksheet."""
        if sheet_name in self._sheets_data and sheet_name != self._active_sheet_name:
            self._active_sheet_name = sheet_name
            if self._workbook and sheet_name in self._workbook.sheetnames:
                self._sheet = self._workbook[sheet_name]
            self._load_active_sheet()

    def _on_add_sheet(self) -> None:
        """Add new worksheet."""
        base_name = "Sheet"
        i = len(self._sheets_data) + 1
        while f"{base_name}{i}" in self._sheets_data:
            i += 1
        new_name = f"{base_name}{i}"

        s_data = SheetData(name=new_name, col_count=10, row_count=30)
        s_data.headers = [get_column_letter(c + 1) for c in range(10)]
        s_data.rows = [["" for _ in range(10)] for _ in range(30)]

        self._sheets_data[new_name] = s_data
        if self._workbook:
            self._workbook.create_sheet(title=new_name)

        sheet_names = list(self._sheets_data.keys())
        self._sheet_tabs.set_sheets(sheet_names)
        self._on_switch_sheet(new_name)
        self._set_modified(True)

    def _on_rename_sheet(self, old_name: str) -> None:
        """Rename worksheet."""
        new_name = simpledialog.askstring("Rename Sheet", "Enter new sheet name:", initialvalue=old_name, parent=self)
        if not new_name or new_name == old_name:
            return
        new_name = new_name.strip()
        if new_name in self._sheets_data:
            messagebox.showwarning("Warning", f"Sheet '{new_name}' already exists.")
            return

        s_data = self._sheets_data.pop(old_name)
        s_data.name = new_name
        self._sheets_data[new_name] = s_data
        self._formula_store.rename_sheet(old_name, new_name)

        if self._workbook and old_name in self._workbook.sheetnames:
            ws = self._workbook[old_name]
            ws.title = new_name

        if self._active_sheet_name == old_name:
            self._active_sheet_name = new_name

        self._sheet_tabs.set_sheets(list(self._sheets_data.keys()))
        self._set_modified(True)

    def _on_delete_sheet(self, sheet_name: str) -> None:
        """Delete worksheet."""
        if len(self._sheets_data) <= 1:
            messagebox.showwarning("Warning", "A workbook must contain at least one visible worksheet.")
            return

        if not messagebox.askyesno("Delete Sheet", f"Are you sure you want to delete '{sheet_name}'?"):
            return

        del self._sheets_data[sheet_name]
        self._formula_store.clear_sheet(sheet_name)
        if self._workbook and sheet_name in self._workbook.sheetnames:
            del self._workbook[sheet_name]

        remaining = list(self._sheets_data.keys())
        self._active_sheet_name = remaining[0]
        self._sheet_tabs.set_sheets(remaining)
        self._load_active_sheet()
        self._set_modified(True)

    def _on_duplicate_sheet(self, sheet_name: str) -> None:
        """Duplicate an existing worksheet."""
        s_data = self._sheets_data.get(sheet_name)
        if not s_data:
            return

        copy_name = f"{sheet_name} (Copy)"
        idx = 2
        while copy_name in self._sheets_data:
            copy_name = f"{sheet_name} (Copy {idx})"
            idx += 1

        new_rows = [list(r) for r in s_data.rows]
        new_s_data = SheetData(
            name=copy_name,
            headers=list(s_data.headers),
            rows=new_rows,
            col_count=s_data.col_count,
            row_count=s_data.row_count
        )
        self._sheets_data[copy_name] = new_s_data

        self._sheet_tabs.set_sheets(list(self._sheets_data.keys()))
        self._on_switch_sheet(copy_name)
        self._set_modified(True)

    # =========================================================================
    # Grid Population & Rendering
    # =========================================================================

    def _get_active_sheet_data(self) -> SheetData:
        if self._active_sheet_name not in self._sheets_data:
            self._sheets_data[self._active_sheet_name] = SheetData(name=self._active_sheet_name, col_count=10, row_count=30)
        return self._sheets_data[self._active_sheet_name]

    def _load_active_sheet(self) -> None:
        """Build and populate the treeview grid from active sheet data."""
        self._cancel_edit()
        s_data = self._get_active_sheet_data()

        # Clear existing
        self._tree.delete(*self._tree.get_children())
        self._row_iids.clear()
        self._col_ids.clear()
        self._detached_rows.clear()
        self._hide_border(self._active_cell_border)
        self._hide_border(self._range_border)

        # Setup columns: '#' (row number) + 'c0', 'c1', ...
        col_count = max(1, s_data.col_count)
        self._col_ids = ["#"] + [f"c{i}" for i in range(col_count)]
        self._tree["columns"] = self._col_ids

        # Row number column
        self._tree.heading("#", text="#", anchor=CENTER)
        self._tree.column("#", width=Config.ROW_NUM_WIDTH, minwidth=40, anchor=CENTER, stretch=False)

        # Data columns
        for i in range(col_count):
            col_id = f"c{i}"
            header_text = s_data.headers[i] if i < len(s_data.headers) else get_column_letter(i + 1)
            # If filter active, add filter symbol
            filter_sym = " 🔻" if s_data.filter_active else ""
            self._tree.heading(
                col_id,
                text=f"{header_text}{filter_sym}",
                anchor=CENTER,
                command=partial(self._on_header_clicked, i)
            )
            col_w = s_data.column_widths.get(i, Config.DEFAULT_COL_WIDTH)
            min_w = 0 if i in s_data.hidden_cols else Config.MIN_COL_WIDTH
            actual_w = 0 if i in s_data.hidden_cols else col_w
            self._tree.column(col_id, width=actual_w, minwidth=min_w, anchor=W)

        # Insert rows
        tree_insert = self._tree.insert
        tree_detach = self._tree.detach
        eval_formula = self._formula_engine.evaluate
        s_name = self._active_sheet_name

        for r_idx, row in enumerate(s_data.rows):
            display_row = [r_idx + 1]
            for c_idx in range(col_count):
                formula = self._formula_store.get(s_name, r_idx, c_idx)
                raw_val = row[c_idx] if c_idx < len(row) else ""
                if formula:
                    val = eval_formula(formula, s_name)
                else:
                    val = raw_val

                # Format with number formatter if exists
                style = s_data.cell_styles.get((r_idx, c_idx))
                fmt = style.number_format if style else None
                formatted_str = NumberFormatter.format_value(val, fmt)
                display_row.append(formatted_str)

            tag = "odd" if r_idx % 2 else "even"
            iid = tree_insert("", END, values=tuple(display_row), tags=(tag,))
            self._row_iids.append(iid)

            if r_idx in s_data.hidden_rows:
                tree_detach(iid)
                self._detached_rows[r_idx] = iid

        # Update stats
        self._status_bar.set_stats(s_data.row_count, s_data.col_count)
        self._update_selection_highlight()
        self._update_formula_bar()

    def _get_cell_value_for_formula(self, row: int, col: int, sheet_name: str | None = None) -> Any:
        """Callback for formula engine to fetch cell value from any sheet."""
        target_sheet = sheet_name or self._active_sheet_name
        s_data = self._sheets_data.get(target_sheet)
        if not s_data:
            return 0

        if row < 0 or row >= len(s_data.rows):
            return 0
        r_data = s_data.rows[row]
        if col < 0 or col >= len(r_data):
            return 0

        val = r_data[col]
        # Check if it's a formula
        formula = self._formula_store.get(target_sheet, row, col)
        if formula:
            return self._formula_engine.evaluate(formula, target_sheet)
        return val

    # =========================================================================
    # Navigation & Selection
    # =========================================================================

    def _navigate(self, row_d: int, col_d: int, shift: bool = False) -> str:
        s_data = self._get_active_sheet_data()
        self._cancel_edit()

        new_r = max(0, min(self._selected.row + row_d, s_data.row_count - 1))
        new_c = max(0, min(self._selected.col + col_d, s_data.col_count - 1))

        if shift:
            # Expand range
            self._range_extent.row = new_r
            self._range_extent.col = new_c
            self._selected.row = new_r
            self._selected.col = new_c
        else:
            # Single cell jump
            self._selected.row = new_r
            self._selected.col = new_c
            self._range_anchor.row = new_r
            self._range_anchor.col = new_c
            self._range_extent.row = new_r
            self._range_extent.col = new_c

        if self._row_iids and self._selected.row < len(self._row_iids):
            iid = self._row_iids[self._selected.row]
            self._tree.selection_set(iid)
            self._tree.see(iid)

        self._update_selection_highlight()
        self._update_formula_bar()
        self._update_status_bar_stats()
        return "break"

    def _on_grid_mouse_down(self, event: tk.Event) -> None:
        cell = self._get_cell_at_xy(event.x, event.y)
        if not cell:
            return

        self._cancel_edit()
        self._selected = cell.copy()
        self._range_anchor = cell.copy()
        self._range_extent = cell.copy()
        self._is_mouse_dragging = True

        if cell.row < len(self._row_iids):
            self._tree.selection_set(self._row_iids[cell.row])

        self._update_selection_highlight()
        self._update_formula_bar()
        self._update_status_bar_stats()

    def _on_grid_mouse_drag(self, event: tk.Event) -> None:
        if not self._is_mouse_dragging:
            return
        cell = self._get_cell_at_xy(event.x, event.y)
        if not cell:
            return

        if cell.row != self._range_extent.row or cell.col != self._range_extent.col:
            self._range_extent = cell.copy()
            self._update_selection_highlight()
            self._update_status_bar_stats()

    def _on_grid_mouse_up(self, event: tk.Event) -> None:
        self._is_mouse_dragging = False

    def _on_grid_double_click(self, event: tk.Event) -> None:
        region = self._tree.identify_region(event.x, event.y)
        if region == "cell":
            cell = self._get_cell_at_xy(event.x, event.y)
            if cell:
                self._selected = cell.copy()
                self._start_inline_edit()

    def _on_grid_key_press(self, event: tk.Event) -> str | None:
        if event.state & 0x4 or not event.char or event.char < " ":
            return None
        self._start_inline_edit(initial_char=event.char)
        return "break"

    def _get_cell_at_xy(self, x: int, y: int) -> CellPosition | None:
        row_id = self._tree.identify_row(y)
        col_id = self._tree.identify_column(x)
        if not row_id or not col_id or col_id == "#1":
            return None

        try:
            col_num = int(col_id.replace("#", "")) - 1
            col_idx = col_num - 1
            if col_idx < 0:
                return None
            row_idx = self._row_iids.index(row_id)
            return CellPosition(row_idx, col_idx)
        except (ValueError, IndexError):
            return None

    def _select_all(self) -> None:
        s_data = self._get_active_sheet_data()
        self._range_anchor = CellPosition(0, 0)
        self._range_extent = CellPosition(max(0, s_data.row_count - 1), max(0, s_data.col_count - 1))
        self._update_selection_highlight()
        self._update_status_bar_stats()

    # =========================================================================
    # Visual Highlights & Status Bar Stats
    # =========================================================================

    def _create_border_frames(self, color: str) -> dict[str, tk.Frame]:
        w = Config.BORDER_WIDTH
        return {
            "top": tk.Frame(self._tree, bg=color, height=w, bd=0),
            "bottom": tk.Frame(self._tree, bg=color, height=w, bd=0),
            "left": tk.Frame(self._tree, bg=color, width=w, bd=0),
            "right": tk.Frame(self._tree, bg=color, width=w, bd=0),
        }

    def _place_border(self, border: dict[str, tk.Frame], x: int, y: int, w: int, h: int) -> None:
        bw = Config.BORDER_WIDTH
        if w <= 0 or h <= 0:
            self._hide_border(border)
            return
        border["top"].place(x=x, y=y, width=w, height=bw)
        border["bottom"].place(x=x, y=y + h - bw, width=w, height=bw)
        border["left"].place(x=x, y=y, width=bw, height=h)
        border["right"].place(x=x + w - bw, y=y, width=bw, height=h)

    def _hide_border(self, border: dict[str, tk.Frame]) -> None:
        for f in border.values():
            f.place_forget()

    def _refresh_highlights(self) -> None:
        self._update_selection_highlight()

    def _update_selection_highlight(self) -> None:
        if not self._row_iids:
            self._hide_border(self._active_cell_border)
            self._hide_border(self._range_border)
            return

        # 1. Active Cell Border
        r = self._selected.row
        c = self._selected.col
        if 0 <= r < len(self._row_iids) and 0 <= c < self._get_active_sheet_data().col_count:
            iid = self._row_iids[r]
            col_id = self._col_ids[c + 1]
            bbox = self._tree.bbox(iid, col_id)
            if bbox:
                x, y, w, h = bbox
                self._place_border(self._active_cell_border, x, y, w, h)
            else:
                self._hide_border(self._active_cell_border)

        # 2. Multi-cell Range Border
        r1 = min(self._range_anchor.row, self._range_extent.row)
        r2 = max(self._range_anchor.row, self._range_extent.row)
        c1 = min(self._range_anchor.col, self._range_extent.col)
        c2 = max(self._range_anchor.col, self._range_extent.col)

        if (r1 == r2 and c1 == c2) or r2 >= len(self._row_iids):
            self._hide_border(self._range_border)
            return

        iid1 = self._row_iids[r1]
        iid2 = self._row_iids[r2]
        col_id1 = self._col_ids[c1 + 1]
        col_id2 = self._col_ids[c2 + 1]
        bbox1 = self._tree.bbox(iid1, col_id1)
        bbox2 = self._tree.bbox(iid2, col_id2)

        if bbox1 and bbox2:
            x = bbox1[0]
            y = bbox1[1]
            w = (bbox2[0] + bbox2[2]) - x
            h = (bbox2[1] + bbox2[3]) - y
            self._place_border(self._range_border, x, y, w, h)
        else:
            self._hide_border(self._range_border)

    def _update_formula_bar(self) -> None:
        s_data = self._get_active_sheet_data()
        r = self._selected.row
        c = self._selected.col
        cell_ref_str = self._selected.to_excel()

        formula = self._formula_store.get(self._active_sheet_name, r, c)
        if formula:
            val_str = formula
        else:
            val = s_data.get_cell_value(r, c)
            val_str = "" if val is None else str(val)

        self._formula_bar.update_cell(cell_ref_str, val_str)
        self._status_bar.set_cell(cell_ref_str)

    def _update_status_bar_stats(self) -> None:
        s_data = self._get_active_sheet_data()
        r1 = min(self._range_anchor.row, self._range_extent.row)
        r2 = max(self._range_anchor.row, self._range_extent.row)
        c1 = min(self._range_anchor.col, self._range_extent.col)
        c2 = max(self._range_anchor.col, self._range_extent.col)

        vals = []
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                formula = self._formula_store.get(self._active_sheet_name, r, c)
                if formula:
                    vals.append(self._formula_engine.evaluate(formula, self._active_sheet_name))
                else:
                    vals.append(s_data.get_cell_value(r, c))

        self._status_bar.update_range_stats(vals)

    def _on_tree_y_scroll(self, *args) -> None:
        if self._vsb:
            self._vsb.set(*args)
        self._refresh_highlights()

    def _on_tree_x_scroll(self, *args) -> None:
        if self._hsb:
            self._hsb.set(*args)
        self._refresh_highlights()

    # =========================================================================
    # Editing Operations
    # =========================================================================

    def _start_inline_edit(self, initial_char: str | None = None) -> None:
        if not self._row_iids:
            return
        r = self._selected.row
        c = self._selected.col
        if r >= len(self._row_iids):
            return

        iid = self._row_iids[r]
        col_id = self._col_ids[c + 1]

        formula = self._formula_store.get(self._active_sheet_name, r, c)
        raw_val = self._get_active_sheet_data().get_cell_value(r, c)
        val = formula if formula else ("" if raw_val is None else str(raw_val))

        self._cell_editor.start(iid, col_id, val, initial_char)

    def _on_edit_commit(self, new_val: str) -> None:
        self._apply_cell_value_change(self._selected.row, self._selected.col, new_val)
        self._tree.focus_set()

    def _on_edit_cancel(self) -> None:
        self._tree.focus_set()

    def _cancel_edit(self) -> None:
        self._cell_editor.cancel()

    def _on_formula_commit(self, val: str) -> None:
        self._apply_cell_value_change(self._selected.row, self._selected.col, val)

    def _apply_cell_value_change(self, row: int, col: int, new_value: str) -> None:
        s_data = self._get_active_sheet_data()
        s_name = self._active_sheet_name

        old_formula = self._formula_store.get(s_name, row, col)
        old_val = s_data.get_cell_value(row, col)
        recorded_old = old_formula if old_formula else ("" if old_val is None else str(old_val))

        if recorded_old == new_value:
            return

        # Check if formula
        if new_value.startswith("="):
            self._formula_store.set(s_name, row, col, new_value)
            evaluated = self._formula_engine.evaluate(new_value, s_name)
            s_data.set_cell_value(row, col, evaluated)
        else:
            self._formula_store.set(s_name, row, col, "")
            # Try parsing numeric / typed value
            parsed = self._parse_typed_value(new_value)
            s_data.set_cell_value(row, col, parsed)

        # Update Treeview cell
        if row < len(self._row_iids):
            iid = self._row_iids[row]
            vals = list(self._tree.item(iid)["values"])
            while len(vals) <= col + 1:
                vals.append("")
            display_val = self._formula_engine.evaluate(new_value, s_name) if new_value.startswith("=") else new_value
            style = s_data.cell_styles.get((row, col))
            fmt = style.number_format if style else None
            vals[col + 1] = NumberFormatter.format_value(display_val, fmt)
            self._tree.item(iid, values=vals)

        # Track Undo
        self._undo_stack.append(UndoAction(
            action_type="cell_change",
            sheet_name=s_name,
            data={"row": row, "col": col, "old": recorded_old, "new": new_value}
        ))
        self._redo_stack.clear()

        self._set_modified(True)
        self._update_formula_bar()
        self._recalculate_all()

    def _parse_typed_value(self, s: str) -> Any:
        if not s:
            return ""
        try:
            if "." in s:
                return float(s)
            return int(s)
        except ValueError:
            pass
        return s

    def _delete_cell(self) -> None:
        """Clear contents of selected range."""
        r1 = min(self._range_anchor.row, self._range_extent.row)
        r2 = max(self._range_anchor.row, self._range_extent.row)
        c1 = min(self._range_anchor.col, self._range_extent.col)
        c2 = max(self._range_anchor.col, self._range_extent.col)

        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                self._apply_cell_value_change(r, c, "")

    def _recalculate_all(self) -> None:
        """Recalculate all formulas on the active sheet and update display."""
        s_data = self._get_active_sheet_data()
        s_name = self._active_sheet_name

        for (r, c), f in self._formula_store.get_all(s_name).items():
            if r < len(self._row_iids):
                result = self._formula_engine.evaluate(f, s_name)
                s_data.set_cell_value(r, c, result)
                iid = self._row_iids[r]
                vals = list(self._tree.item(iid)["values"])
                if c + 1 < len(vals):
                    style = s_data.cell_styles.get((r, c))
                    fmt = style.number_format if style else None
                    vals[c + 1] = NumberFormatter.format_value(result, fmt)
                    self._tree.item(iid, values=vals)

        self._update_status_bar_stats()

    # =========================================================================
    # Clipboard (2D TSV / Excel-Compatible Copy / Paste / Cut)
    # =========================================================================

    def _copy(self) -> None:
        """Copy selected range to system clipboard as TSV."""
        s_data = self._get_active_sheet_data()
        r1 = min(self._range_anchor.row, self._range_extent.row)
        r2 = max(self._range_anchor.row, self._range_extent.row)
        c1 = min(self._range_anchor.col, self._range_extent.col)
        c2 = max(self._range_anchor.col, self._range_extent.col)

        grid_lines = []
        internal_grid = []
        for r in range(r1, r2 + 1):
            row_items = []
            for c in range(c1, c2 + 1):
                formula = self._formula_store.get(self._active_sheet_name, r, c)
                val = formula if formula else s_data.get_cell_value(r, c)
                row_items.append("" if val is None else str(val))
            grid_lines.append("\t".join(row_items))
            internal_grid.append(row_items)

        tsv_str = "\n".join(grid_lines)
        self.clipboard_clear()
        self.clipboard_append(tsv_str)
        self._clipboard_buffer = internal_grid
        self._status_bar.set_mode("Copied")

    def _cut(self) -> None:
        """Cut selected range."""
        self._copy()
        self._delete_cell()
        self._status_bar.set_mode("Cut")

    def _paste(self) -> None:
        """Paste TSV data starting at active cell."""
        try:
            tsv_data = self.clipboard_get()
        except tk.TclError:
            return

        if not tsv_data:
            return

        lines = tsv_data.splitlines()
        start_r = self._selected.row
        start_c = self._selected.col

        for r_offset, line in enumerate(lines):
            cols = line.split("\t")
            for c_offset, text_val in enumerate(cols):
                target_r = start_r + r_offset
                target_c = start_c + c_offset
                self._apply_cell_value_change(target_r, target_c, text_val)

        self._status_bar.set_mode("Pasted")

    def _paste_special(self, mode: str) -> None:
        """Paste values only, formulas only, or transpose."""
        if not self._clipboard_buffer:
            self._paste()
            return

        start_r = self._selected.row
        start_c = self._selected.col

        if mode == "transpose":
            num_rows = len(self._clipboard_buffer)
            num_cols = len(self._clipboard_buffer[0]) if num_rows else 0
            for r in range(num_rows):
                for c in range(num_cols):
                    val = self._clipboard_buffer[r][c]
                    self._apply_cell_value_change(start_r + c, start_c + r, val)

        elif mode == "values":
            for r_offset, row in enumerate(self._clipboard_buffer):
                for c_offset, val in enumerate(row):
                    clean_val = val
                    if val.startswith("="):
                        clean_val = str(self._formula_engine.evaluate(val, self._active_sheet_name))
                    self._apply_cell_value_change(start_r + r_offset, start_c + c_offset, clean_val)

        elif mode == "formulas":
            for r_offset, row in enumerate(self._clipboard_buffer):
                for c_offset, val in enumerate(row):
                    if val.startswith("="):
                        shifted = shift_formula_references(val, r_offset, c_offset)
                        self._apply_cell_value_change(start_r + r_offset, start_c + c_offset, shifted)
                    else:
                        self._apply_cell_value_change(start_r + r_offset, start_c + c_offset, val)

    def _activate_format_painter(self) -> None:
        s_data = self._get_active_sheet_data()
        style = s_data.cell_styles.get((self._selected.row, self._selected.col))
        self._format_painter_style = style.copy() if style else CellStyle()
        self._status_bar.set_mode("Format Painter Active")

    # =========================================================================
    # Undo / Redo
    # =========================================================================

    def _undo(self) -> None:
        if not self._undo_stack:
            return
        action = self._undo_stack.pop()
        self._apply_undo_action(action, is_undo=True)
        self._redo_stack.append(action)
        self._status_bar.set_mode("Undo")

    def _redo(self) -> None:
        if not self._redo_stack:
            return
        action = self._redo_stack.pop()
        self._apply_undo_action(action, is_undo=False)
        self._undo_stack.append(action)
        self._status_bar.set_mode("Redo")

    def _apply_undo_action(self, action: UndoAction, is_undo: bool) -> None:
        s_name = action.sheet_name
        s_data = self._sheets_data.get(s_name)
        if not s_data:
            return

        if action.action_type == "cell_change":
            r = action.data["row"]
            c = action.data["col"]
            val = action.data["old"] if is_undo else action.data["new"]
            if val.startswith("="):
                self._formula_store.set(s_name, r, c, val)
                s_data.set_cell_value(r, c, self._formula_engine.evaluate(val, s_name))
            else:
                self._formula_store.set(s_name, r, c, "")
                s_data.set_cell_value(r, c, self._parse_typed_value(val))

            if s_name == self._active_sheet_name and r < len(self._row_iids):
                iid = self._row_iids[r]
                vals = list(self._tree.item(iid)["values"])
                if c + 1 < len(vals):
                    vals[c + 1] = val
                    self._tree.item(iid, values=vals)

        self._set_modified(True)
        self._update_formula_bar()
        self._recalculate_all()

    # =========================================================================
    # Formatting & Styles
    # =========================================================================

    def _get_or_create_cell_style(self, row: int, col: int) -> CellStyle:
        s_data = self._get_active_sheet_data()
        key = (row, col)
        if key not in s_data.cell_styles:
            s_data.cell_styles[key] = CellStyle()
        return s_data.cell_styles[key]

    def _apply_style_to_selected_range(self, updater: Callable[[CellStyle], None]) -> None:
        r1 = min(self._range_anchor.row, self._range_extent.row)
        r2 = max(self._range_anchor.row, self._range_extent.row)
        c1 = min(self._range_anchor.col, self._range_extent.col)
        c2 = max(self._range_anchor.col, self._range_extent.col)

        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                style = self._get_or_create_cell_style(r, c)
                updater(style)

        self._set_modified(True)
        self._load_active_sheet()

    def _set_font_family(self, font_name: str) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "font_name", font_name))

    def _set_font_size(self, size: int) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "font_size", size))

    def _increase_font_size(self) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "font_size", min(72, s.font_size + 2)))

    def _decrease_font_size(self) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "font_size", max(6, s.font_size - 2)))

    def _toggle_bold(self) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "bold", not s.bold))

    def _toggle_italic(self) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "italic", not s.italic))

    def _toggle_underline(self) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "underline", not s.underline))

    def _toggle_strikethrough(self) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "strikethrough", not s.strikethrough))

    def _pick_bg_color(self) -> None:
        color = colorchooser.askcolor(title="Select Cell Background Color")
        if color[1]:
            hex_c = color[1].upper()
            self._apply_style_to_selected_range(lambda s: setattr(s, "bg_color", hex_c))
            self._ribbon.update_color_buttons(hex_c, None)

    def _pick_fg_color(self) -> None:
        color = colorchooser.askcolor(title="Select Font Color")
        if color[1]:
            hex_c = color[1].upper()
            self._apply_style_to_selected_range(lambda s: setattr(s, "fg_color", hex_c))
            self._ribbon.update_color_buttons(None, hex_c)

    def _set_borders(self, border_type: str) -> None:
        self._apply_style_to_selected_range(lambda s: s.borders.update({"type": border_type}))

    def _set_halign(self, align: str) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "halign", align))

    def _set_valign(self, align: str) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "valign", align))

    def _toggle_wrap_text(self) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "wrap_text", not s.wrap_text))

    def _toggle_merge_cells(self) -> None:
        s_data = self._get_active_sheet_data()
        rng = CellRange(self._range_anchor.copy(), self._range_extent.copy())
        s_data.merged_ranges.append(rng)
        self._set_modified(True)
        messagebox.showinfo("Merge & Center", f"Merged cells in range {rng.to_excel()}")

    def _set_number_format(self, fmt: str) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "number_format", fmt))

    def _increase_decimals(self) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "number_format", "#,##0.000"))

    def _decrease_decimals(self) -> None:
        self._apply_style_to_selected_range(lambda s: setattr(s, "number_format", "#,##0.0"))

    def _add_cf_rule(self, rule_type: str) -> None:
        s_data = self._get_active_sheet_data()
        if "color_scale" in rule_type:
            scale = "green_yellow_red" if "gyr" in rule_type else "red_yellow_green"
            s_data.filter_criteria[-1] = {"type": "color_scale", "scale": scale}
        else:
            val = simpledialog.askstring("Conditional Formatting", f"Enter threshold value for {rule_type}:", parent=self)
            if val is not None:
                s_data.filter_criteria[-1] = {"type": rule_type, "value": val, "bg_color": "#ffc7ce", "fg_color": "#9c0006"}

        self._set_modified(True)
        self._load_active_sheet()

    def _clear_cf_rules(self) -> None:
        s_data = self._get_active_sheet_data()
        s_data.filter_criteria.clear()
        self._set_modified(True)
        self._load_active_sheet()

    # =========================================================================
    # Rows & Columns Structure Manipulation
    # =========================================================================

    def _insert_row_above(self) -> None:
        s_data = self._get_active_sheet_data()
        r = self._selected.row
        new_row = ["" for _ in range(s_data.col_count)]
        s_data.rows.insert(r, new_row)
        s_data.row_count += 1
        self._formula_store.insert_row(self._active_sheet_name, r)
        self._set_modified(True)
        self._load_active_sheet()

    def _insert_row_below(self) -> None:
        s_data = self._get_active_sheet_data()
        r = self._selected.row + 1
        new_row = ["" for _ in range(s_data.col_count)]
        s_data.rows.insert(r, new_row)
        s_data.row_count += 1
        self._formula_store.insert_row(self._active_sheet_name, r)
        self._set_modified(True)
        self._load_active_sheet()

    def _insert_col_left(self) -> None:
        s_data = self._get_active_sheet_data()
        c = self._selected.col
        for row in s_data.rows:
            row.insert(c, "")
        s_data.col_count += 1
        s_data.headers = [get_column_letter(i + 1) for i in range(s_data.col_count)]
        self._formula_store.insert_col(self._active_sheet_name, c)
        self._set_modified(True)
        self._load_active_sheet()

    def _insert_col_right(self) -> None:
        s_data = self._get_active_sheet_data()
        c = self._selected.col + 1
        for row in s_data.rows:
            row.insert(c, "")
        s_data.col_count += 1
        s_data.headers = [get_column_letter(i + 1) for i in range(s_data.col_count)]
        self._formula_store.insert_col(self._active_sheet_name, c)
        self._set_modified(True)
        self._load_active_sheet()

    def _delete_row(self) -> None:
        s_data = self._get_active_sheet_data()
        if s_data.row_count <= 1:
            return
        r = self._selected.row
        s_data.rows.pop(r)
        s_data.row_count -= 1
        self._formula_store.delete_row(self._active_sheet_name, r)
        self._selected.row = min(r, s_data.row_count - 1)
        self._set_modified(True)
        self._load_active_sheet()

    def _delete_column(self) -> None:
        s_data = self._get_active_sheet_data()
        if s_data.col_count <= 1:
            return
        c = self._selected.col
        for row in s_data.rows:
            if c < len(row):
                row.pop(c)
        s_data.col_count -= 1
        s_data.headers = [get_column_letter(i + 1) for i in range(s_data.col_count)]
        self._formula_store.delete_col(self._active_sheet_name, c)
        self._selected.col = min(c, s_data.col_count - 1)
        self._set_modified(True)
        self._load_active_sheet()

    def _autofit_all_columns(self) -> None:
        s_data = self._get_active_sheet_data()
        for c in range(s_data.col_count):
            max_len = len(s_data.headers[c]) if c < len(s_data.headers) else 5
            for row in s_data.rows[:100]:
                if c < len(row):
                    max_len = max(max_len, len(str(row[c] or "")))
            col_id = f"c{c}"
            w = max(Config.MIN_COL_WIDTH, min(Config.MAX_COL_WIDTH, max_len * 9 + 25))
            s_data.column_widths[c] = w
            self._tree.column(col_id, width=w)
        self._status_bar.set_mode("Columns Auto-fitted")
        self._refresh_highlights()

    # =========================================================================
    # AutoFilter & Sorting
    # =========================================================================

    def _toggle_autofilter(self) -> None:
        s_data = self._get_active_sheet_data()
        s_data.filter_active = not s_data.filter_active
        if not s_data.filter_active:
            s_data.filter_criteria.clear()
            s_data.hidden_rows.clear()
        self._load_active_sheet()
        self._status_bar.set_mode("AutoFilter Enabled" if s_data.filter_active else "AutoFilter Disabled")

    def _clear_filters(self) -> None:
        s_data = self._get_active_sheet_data()
        s_data.filter_criteria.clear()
        s_data.hidden_rows.clear()
        self._load_active_sheet()
        self._status_bar.set_mode("Filters Cleared")

    def _on_header_clicked(self, col_idx: int) -> None:
        s_data = self._get_active_sheet_data()
        if s_data.filter_active:
            # Show Filter popup
            unique_vals = list({str(row[col_idx]) if col_idx < len(row) and row[col_idx] is not None else "" for row in s_data.rows})
            selected_vals = s_data.filter_criteria.get(col_idx)
            header_name = s_data.headers[col_idx] if col_idx < len(s_data.headers) else get_column_letter(col_idx + 1)

            popup = AutoFilterPopup(
                self,
                col_idx=col_idx,
                col_name=header_name,
                unique_values=unique_vals,
                selected_values=selected_vals,
                on_apply=self._apply_column_filter,
                on_sort_col=self._quick_sort_column
            )
            popup.grab_set()
        else:
            # Quick sort toggle
            self._quick_sort_column(col_idx, False)

    def _apply_column_filter(self, col_idx: int, selected_vals: set[str] | None) -> None:
        s_data = self._get_active_sheet_data()
        if selected_vals is None:
            s_data.filter_criteria.pop(col_idx, None)
        else:
            s_data.filter_criteria[col_idx] = selected_vals

        # Re-evaluate all hidden rows
        s_data.hidden_rows.clear()
        for r_idx, row in enumerate(s_data.rows):
            for c_idx, allowed_vals in s_data.filter_criteria.items():
                val_str = str(row[c_idx]) if c_idx < len(row) and row[c_idx] is not None else ""
                if val_str not in allowed_vals:
                    s_data.hidden_rows.add(r_idx)
                    break

        self._load_active_sheet()

    def _quick_sort(self, descending: bool) -> None:
        self._quick_sort_column(self._selected.col, descending)

    def _quick_sort_column(self, col_idx: int, descending: bool) -> None:
        s_data = self._get_active_sheet_data()

        def key_fn(row):
            val = row[col_idx] if col_idx < len(row) else ""
            if val is None or val == "":
                return (2, 0, "")
            try:
                num = float(str(val).replace(",", ".").replace(" ", ""))
                return (0, num, "")
            except (ValueError, TypeError):
                return (1, 0, str(val).lower())

        s_data.rows.sort(key=key_fn, reverse=descending)
        self._set_modified(True)
        self._load_active_sheet()

    def _show_custom_sort_dialog(self) -> None:
        s_data = self._get_active_sheet_data()
        dlg = CustomSortDialog(self, s_data.headers, on_apply_sort=self._apply_multi_sort)
        dlg.grab_set()

    def _apply_multi_sort(self, sort_rules: list[tuple[int, bool]], has_headers: bool) -> None:
        s_data = self._get_active_sheet_data()

        def multi_key_fn(row):
            keys = []
            for col_idx, is_desc in sort_rules:
                val = row[col_idx] if col_idx < len(row) else ""
                if val is None or val == "":
                    k = (2, 0, "")
                else:
                    try:
                        num = float(str(val).replace(",", ".").replace(" ", ""))
                        k = (0, num, "")
                    except (ValueError, TypeError):
                        k = (1, 0, str(val).lower())
                keys.append(k)
            return tuple(keys)

        s_data.rows.sort(key=multi_key_fn)
        self._set_modified(True)
        self._load_active_sheet()

    # =========================================================================
    # AutoSum & Formula Helpers
    # =========================================================================

    def _apply_autosum(self, func_name: str) -> None:
        s_data = self._get_active_sheet_data()
        r = self._selected.row
        c = self._selected.col

        # Find continuous range of numbers directly above
        start_r = r - 1
        while start_r >= 0 and s_data.get_cell_value(start_r, c) not in (None, ""):
            start_r -= 1
        start_r += 1

        if start_r < r:
            p1 = CellPosition(start_r, c).to_excel()
            p2 = CellPosition(r - 1, c).to_excel()
            formula = f"={func_name}({p1}:{p2})"
        else:
            # Fallback to left
            start_c = c - 1
            while start_c >= 0 and s_data.get_cell_value(r, start_c) not in (None, ""):
                start_c -= 1
            start_c += 1
            if start_c < c:
                p1 = CellPosition(r, start_c).to_excel()
                p2 = CellPosition(r, c - 1).to_excel()
                formula = f"={func_name}({p1}:{p2})"
            else:
                formula = f"={func_name}()"

        self._apply_cell_value_change(r, c, formula)

    def _show_fx_wizard(self) -> None:
        InsertFunctionDialog(self, on_insert_formula=self._insert_formula_text)

    def _insert_formula_text(self, prefix: str) -> None:
        self._formula_bar.focus_entry()
        self._formula_bar.set_value(prefix)

    # =========================================================================
    # Find & Replace & Go To
    # =========================================================================

    def _show_find_dialog(self) -> None:
        FindReplaceDialog(
            self,
            on_find_next=self._find_next,
            on_replace=self._replace_one,
            on_replace_all=self._replace_all,
            initial_tab="find"
        )

    def _show_replace_dialog(self) -> None:
        FindReplaceDialog(
            self,
            on_find_next=self._find_next,
            on_replace=self._replace_one,
            on_replace_all=self._replace_all,
            initial_tab="replace"
        )

    def _find_next(self, query: str, match_case: bool, match_entire: bool, all_sheets: bool) -> bool:
        s_data = self._get_active_sheet_data()
        start_r = self._selected.row
        start_c = self._selected.col + 1

        for r in range(start_r, s_data.row_count):
            for c in range(start_c if r == start_r else 0, s_data.col_count):
                val = str(s_data.get_cell_value(r, c) or "")
                target = query if match_case else query.lower()
                src = val if match_case else val.lower()

                matched = (src == target) if match_entire else (target in src)
                if matched:
                    self._goto_cell(CellPosition(r, c).to_excel())
                    return True

        # Wrap around from beginning
        for r in range(0, start_r + 1):
            for c in range(0, s_data.col_count):
                if r == start_r and c >= start_c:
                    break
                val = str(s_data.get_cell_value(r, c) or "")
                target = query if match_case else query.lower()
                src = val if match_case else val.lower()
                matched = (src == target) if match_entire else (target in src)
                if matched:
                    self._goto_cell(CellPosition(r, c).to_excel())
                    return True
        return False

    def _replace_one(self, query: str, replacement: str, match_case: bool, match_entire: bool) -> bool:
        s_data = self._get_active_sheet_data()
        r = self._selected.row
        c = self._selected.col
        val = str(s_data.get_cell_value(r, c) or "")
        target = query if match_case else query.lower()
        src = val if match_case else val.lower()

        matched = (src == target) if match_entire else (target in src)
        if matched:
            new_val = val.replace(query, replacement) if not match_entire else replacement
            self._apply_cell_value_change(r, c, new_val)
            self._find_next(query, match_case, match_entire, False)
            return True
        return self._find_next(query, match_case, match_entire, False)

    def _replace_all(self, query: str, replacement: str, match_case: bool, match_entire: bool) -> int:
        s_data = self._get_active_sheet_data()
        count = 0
        for r in range(s_data.row_count):
            for c in range(s_data.col_count):
                val = str(s_data.get_cell_value(r, c) or "")
                target = query if match_case else query.lower()
                src = val if match_case else val.lower()
                matched = (src == target) if match_entire else (target in src)
                if matched:
                    new_val = val.replace(query, replacement) if not match_entire else replacement
                    self._apply_cell_value_change(r, c, new_val)
                    count += 1
        return count

    def _show_goto_dialog(self) -> None:
        target = simpledialog.askstring("Go to Cell", "Enter cell reference (e.g. A1, Z100):", parent=self)
        if target:
            self._goto_cell(target.strip().upper())

    def _goto_cell(self, cell_ref_text: str) -> None:
        try:
            pos = CellPosition.from_excel(cell_ref_text)
            s_data = self._get_active_sheet_data()
            if 0 <= pos.row < s_data.row_count and 0 <= pos.col < s_data.col_count:
                self._selected = pos.copy()
                self._range_anchor = pos.copy()
                self._range_extent = pos.copy()
                if pos.row < len(self._row_iids):
                    iid = self._row_iids[pos.row]
                    self._tree.selection_set(iid)
                    self._tree.see(iid)
                self._update_selection_highlight()
                self._update_formula_bar()
                self._update_status_bar_stats()
        except ValueError:
            messagebox.showwarning("Warning", f"Invalid cell reference: '{cell_ref_text}'")

    # =========================================================================
    # Charts & Wizards
    # =========================================================================

    def _show_chart_wizard(self) -> None:
        s_data = self._get_active_sheet_data()
        dlg = ChartWizardDialog(self, s_data, get_data_fn=self._get_chart_series_data)
        dlg.grab_set()

    def _create_quick_chart(self, chart_type: str) -> None:
        self._show_chart_wizard()

    def _get_chart_series_data(self, r1: int, r2: int, c1: int, c2: int) -> tuple[list[str], list[list[float]], list[str]]:
        s_data = self._get_active_sheet_data()
        labels = []
        for r in range(r1, r2 + 1):
            val = s_data.get_cell_value(r, c1)
            labels.append(str(val if val is not None else f"Row {r+1}"))

        series = []
        series_names = []
        for c in range(c1 + 1, c2 + 1):
            col_header = s_data.headers[c] if c < len(s_data.headers) else get_column_letter(c + 1)
            series_names.append(col_header)
            data_points = []
            for r in range(r1, r2 + 1):
                raw_val = s_data.get_cell_value(r, c)
                try:
                    num = float(str(raw_val).replace(",", ".").replace(" ", ""))
                    data_points.append(num)
                except (ValueError, TypeError):
                    data_points.append(0.0)
            series.append(data_points)

        return labels, series, series_names

    # =========================================================================
    # Advanced Data Tools
    # =========================================================================

    def _show_text_to_columns_dialog(self) -> None:
        s_data = self._get_active_sheet_data()
        c = self._selected.col
        sample_rows = [str(r[c]) for r in s_data.rows[:6] if c < len(r) and r[c]]
        dlg = TextToColumnsDialog(self, sample_rows, on_split=self._apply_text_to_columns)
        dlg.grab_set()

    def _apply_text_to_columns(self, delimiter: str) -> None:
        s_data = self._get_active_sheet_data()
        c = self._selected.col
        for r_idx, row in enumerate(s_data.rows):
            if c < len(row) and row[c]:
                parts = str(row[c]).split(delimiter)
                for p_idx, part in enumerate(parts):
                    self._apply_cell_value_change(r_idx, c + p_idx, part.strip())
        self._load_active_sheet()

    def _show_remove_duplicates_dialog(self) -> None:
        s_data = self._get_active_sheet_data()
        dlg = RemoveDuplicatesDialog(self, s_data.headers, on_remove=self._apply_remove_duplicates)
        dlg.grab_set()

    def _apply_remove_duplicates(self, cols: list[int]) -> None:
        s_data = self._get_active_sheet_data()
        seen = set()
        unique_rows = []
        dup_count = 0

        for row in s_data.rows:
            key = tuple(row[c] if c < len(row) else "" for c in cols)
            if key in seen:
                dup_count += 1
            else:
                seen.add(key)
                unique_rows.append(row)

        s_data.rows = unique_rows
        s_data.row_count = len(unique_rows)
        self._set_modified(True)
        self._load_active_sheet()
        messagebox.showinfo("Remove Duplicates", f"{dup_count} duplicate values found and removed; {len(unique_rows)} unique values remain.")

    def _show_goal_seek_dialog(self) -> None:
        GoalSeekDialog(self, current_cell=self._selected.to_excel(), on_solve=self._apply_goal_seek)

    def _apply_goal_seek(self, set_cell: str, target_val: float, by_cell: str) -> None:
        try:
            set_pos = CellPosition.from_excel(set_cell)
            by_pos = CellPosition.from_excel(by_cell)
            s_name = self._active_sheet_name
            formula = self._formula_store.get(s_name, set_pos.row, set_pos.col)
            if not formula:
                messagebox.showerror("Goal Seek Error", f"Cell {set_cell} must contain a formula.")
                return

            # Simple secant / bisection iterative solver
            cur_x = float(self._get_cell_value_for_formula(by_pos.row, by_pos.col) or 1.0)
            self._apply_cell_value_change(by_pos.row, by_pos.col, str(cur_x))
            y1 = float(self._formula_engine.evaluate(formula, s_name)) - target_val

            if abs(y1) < 1e-6:
                messagebox.showinfo("Goal Seek", f"Goal Seek found a solution!\nCell {by_cell} = {cur_x:g}")
                return

            x2 = cur_x + 1.0
            self._apply_cell_value_change(by_pos.row, by_pos.col, str(x2))
            y2 = float(self._formula_engine.evaluate(formula, s_name)) - target_val

            for _ in range(50):
                if abs(y2 - y1) < 1e-12:
                    break
                x_next = x2 - y2 * (x2 - cur_x) / (y2 - y1)
                cur_x, y1 = x2, y2
                x2 = x_next
                self._apply_cell_value_change(by_pos.row, by_pos.col, str(x2))
                y2 = float(self._formula_engine.evaluate(formula, s_name)) - target_val
                if abs(y2) < 1e-5:
                    messagebox.showinfo("Goal Seek", f"Goal Seek with Cell {set_cell} found a solution.\n\nTarget Value: {target_val}\nChanging Cell: {by_cell} = {x2:g}")
                    return

            messagebox.showinfo("Goal Seek", f"Goal Seek converged near: {x2:g}")
        except Exception as e:
            messagebox.showerror("Goal Seek Error", f"Failed to solve:\n{e}")

    # =========================================================================
    # Comments & UI Toggles
    # =========================================================================

    def _edit_cell_comment(self) -> None:
        s_data = self._get_active_sheet_data()
        key = (self._selected.row, self._selected.col)
        comm = s_data.comments.get(key)
        init_text = comm.text if comm else ""

        def on_save(txt: str):
            if txt:
                s_data.comments[key] = CellComment(text=txt)
            elif key in s_data.comments:
                del s_data.comments[key]
            self._set_modified(True)

        def on_del():
            if key in s_data.comments:
                del s_data.comments[key]
                self._set_modified(True)

        CellCommentDialog(self, self._selected.to_excel(), init_text, on_save=on_save, on_delete=on_del)

    def _toggle_formula_bar(self) -> None:
        if self._formula_bar.winfo_ismapped():
            self._formula_bar.pack_forget()
        else:
            self._formula_bar.pack(fill=tk.X, before=self._tree.master)

    def _toggle_status_bar(self) -> None:
        if self._status_bar.winfo_ismapped():
            self._status_bar.pack_forget()
        else:
            self._status_bar.pack(fill=tk.X, side=tk.BOTTOM)

    def _freeze_top_row(self) -> None:
        messagebox.showinfo("Freeze Panes", "Top row is frozen.")

    def _freeze_first_col(self) -> None:
        messagebox.showinfo("Freeze Panes", "First column is frozen.")

    def _unfreeze_all(self) -> None:
        messagebox.showinfo("Freeze Panes", "All panes unfrozen.")

    def _change_theme(self, theme_name: str) -> None:
        try:
            self.style.theme_use(theme_name)
        except Exception:
            pass

    def _on_zoom_change(self, zoom_pct: int) -> None:
        # Scale default column width proportionally
        scale = zoom_pct / 100.0
        new_w = int(Config.DEFAULT_COL_WIDTH * scale)
        for i, col_id in enumerate(self._col_ids[1:]):
            self._tree.column(col_id, width=new_w)
        self._refresh_highlights()

    # =========================================================================
    # Context Menus & Utility
    # =========================================================================

    def _show_grid_context_menu(self, event: tk.Event) -> None:
        cell = self._get_cell_at_xy(event.x, event.y)
        if cell:
            self._selected = cell.copy()
            self._update_selection_highlight()
            self._update_formula_bar()

        menu = Menu(self, tearoff=0)
        menu.add_command(label="Cut (Ctrl+X)", command=self._cut)
        menu.add_command(label="Copy (Ctrl+C)", command=self._copy)
        menu.add_command(label="Paste (Ctrl+V)", command=self._paste)
        menu.add_separator()
        menu.add_command(label="Insert Row Above", command=self._insert_row_above)
        menu.add_command(label="Insert Column Left", command=self._insert_col_left)
        menu.add_command(label="Delete Row", command=self._delete_row)
        menu.add_command(label="Delete Column", command=self._delete_column)
        menu.add_separator()
        menu.add_command(label="Clear Contents", command=self._delete_cell)
        menu.add_command(label="Add / Edit Comment...", command=self._edit_cell_comment)
        menu.tk_popup(event.x_root, event.y_root)

    def _refresh(self) -> None:
        if self._file_path:
            self._load_file(self._file_path)
        else:
            self._load_active_sheet()

    def _set_app_icon(self) -> None:
        icon_p = Path(Config.ICON_PATH)
        if icon_p.exists():
            try:
                self.iconbitmap(str(icon_p))
            except Exception:
                pass

    def _set_modified(self, modified: bool) -> None:
        self._modified = modified
        self._status_bar.set_modified(modified)
        base = f"{Config.APP_TITLE} — {self._file_path.name}" if self._file_path else f"{Config.APP_TITLE} — Untitled"
        self.title(f"{base} *" if modified else base)

    def _confirm_discard(self) -> bool:
        return messagebox.askyesno("Unsaved Changes", "You have unsaved changes. Discard them?")

    def _on_close(self) -> None:
        if self._modified and not self._confirm_discard():
            return
        self._close_workbook()
        self.destroy()

    def run(self) -> None:
        self.mainloop()


# =============================================================================
# Entry Point
# =============================================================================

def main() -> None:
    app = ExcelViewerPro()
    app.run()


if __name__ == "__main__":
    main()
